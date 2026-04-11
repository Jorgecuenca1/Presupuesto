from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Q
from django.http import HttpResponse
from decimal import Decimal, InvalidOperation
import openpyxl

from core.models import ParametrosSistema, TablaConcejoPersoneria, VigenciaFutura
from .models import (
    RubroGasto, SeccionGasto, FuenteFinanciacion, EjecucionGasto, TipoGasto,
    CifraHistoricaGasto, ServicioDeuda, CostoPersonal,
    ContratoCredito, PagareCredito, AmortizacionPagare,
)
from .forms import (
    RubroGastoForm, SeccionGastoForm, FuenteFinanciacionForm,
    EjecucionGastoForm, ImportarGastosExcelForm,
    CifraHistoricaGastoForm, CostoPersonalForm,
    ContratoCreditoForm, PagareCreditoForm, AmortizacionPagareForm,
)


def _vigencia():
    p = ParametrosSistema.objects.filter(activo=True).first()
    return p.vigencia if p else 2026


def _safe_decimal(val):
    if val is None:
        return Decimal('0')
    try:
        return Decimal(str(val).replace(',', '').replace('$', '').strip())
    except (InvalidOperation, ValueError, TypeError):
        return Decimal('0')


# ─── SECCIONES ────────────────────────────────────────────────────
@login_required
def secciones_list(request):
    vigencia = _vigencia()
    secciones = SeccionGasto.objects.filter(vigencia=vigencia)
    form = SeccionGastoForm(initial={'vigencia': vigencia})
    return render(request, 'gastos/secciones_list.html', {
        'secciones': secciones, 'form': form, 'vigencia': vigencia,
    })


@login_required
def seccion_guardar(request):
    if request.method == 'POST':
        pk = request.POST.get('pk')
        instance = get_object_or_404(SeccionGasto, pk=pk) if pk else None
        form = SeccionGastoForm(request.POST, instance=instance)
        if form.is_valid():
            form.save()
            messages.success(request, 'Sección guardada')
    return redirect('gastos_secciones')


@login_required
def seccion_eliminar(request, pk):
    get_object_or_404(SeccionGasto, pk=pk).delete()
    messages.success(request, 'Sección eliminada')
    return redirect('gastos_secciones')


# ─── FUENTES ──────────────────────────────────────────────────────
@login_required
def fuentes_list(request):
    vigencia = _vigencia()
    fuentes = FuenteFinanciacion.objects.filter(vigencia=vigencia)
    form = FuenteFinanciacionForm(initial={'vigencia': vigencia})
    return render(request, 'gastos/fuentes_list.html', {
        'fuentes': fuentes, 'form': form, 'vigencia': vigencia,
    })


@login_required
def fuente_guardar(request):
    if request.method == 'POST':
        pk = request.POST.get('pk')
        instance = get_object_or_404(FuenteFinanciacion, pk=pk) if pk else None
        form = FuenteFinanciacionForm(request.POST, instance=instance)
        if form.is_valid():
            form.save()
            messages.success(request, 'Fuente guardada')
    return redirect('gastos_fuentes')


@login_required
def fuente_eliminar(request, pk):
    get_object_or_404(FuenteFinanciacion, pk=pk).delete()
    messages.success(request, 'Fuente eliminada')
    return redirect('gastos_fuentes')


# ─── RUBROS DE GASTO ─────────────────────────────────────────────
@login_required
def rubros_gasto_list(request):
    vigencia = _vigencia()
    rubros = RubroGasto.objects.filter(vigencia=vigencia)
    total = rubros.filter(es_titulo=False).aggregate(t=Sum('valor_apropiacion'))['t'] or 0
    return render(request, 'gastos/rubros_gasto_list.html', {
        'rubros': rubros, 'vigencia': vigencia, 'total': total,
    })


@login_required
def rubro_gasto_crear(request):
    vigencia = _vigencia()
    if request.method == 'POST':
        form = RubroGastoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Rubro de gasto creado')
            return redirect('rubros_gasto_list')
    else:
        form = RubroGastoForm(initial={'vigencia': vigencia})
    return render(request, 'gastos/rubro_gasto_form.html', {'form': form, 'titulo': 'Crear Rubro de Gasto'})


@login_required
def rubro_gasto_editar(request, pk):
    rubro = get_object_or_404(RubroGasto, pk=pk)
    if request.method == 'POST':
        form = RubroGastoForm(request.POST, instance=rubro)
        if form.is_valid():
            form.save()
            messages.success(request, 'Rubro actualizado')
            return redirect('rubros_gasto_list')
    else:
        form = RubroGastoForm(instance=rubro)
    return render(request, 'gastos/rubro_gasto_form.html', {'form': form, 'titulo': 'Editar Rubro de Gasto'})


@login_required
def rubro_gasto_eliminar(request, pk):
    get_object_or_404(RubroGasto, pk=pk).delete()
    messages.success(request, 'Rubro eliminado')
    return redirect('rubros_gasto_list')


# ─── RECALCULAR TÍTULOS ──────────────────────────────────────────
@login_required
def recalcular_gastos(request):
    vigencia = _vigencia()
    if request.method == 'POST':
        titulos = RubroGasto.objects.filter(vigencia=vigencia, es_titulo=True).order_by('-nivel')
        for titulo in titulos:
            titulo.calcular_hijos()
        messages.success(request, 'Rubros de gasto recalculados correctamente')
    return redirect('reporte_gastos')


# ─── IMPORTAR ANEXO 2 (PRESUPUESTO DE GASTOS) ────────────────────
@login_required
def importar_anexo2(request):
    if request.method == 'POST':
        form = ImportarGastosExcelForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES['archivo']
            vigencia = _vigencia()
            try:
                wb = openpyxl.load_workbook(archivo, data_only=True)
                # Buscar la hoja principal
                sheet_name = None
                for name in wb.sheetnames:
                    if 'ANEXO' in name.upper() or 'DETALLE' in name.upper() or 'GASTO' in name.upper():
                        sheet_name = name
                        break
                if not sheet_name:
                    sheet_name = wb.sheetnames[0]
                ws = wb[sheet_name]

                # Detectar columnas por encabezados
                header_row = None
                col_map = {}
                for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=False), 1):
                    for cell in row:
                        val = str(cell.value or '').upper().strip()
                        if 'SECCION' in val and 'NOMBRE' not in val and 'INVERSION' not in val:
                            col_map['seccion'] = cell.column - 1
                            header_row = row_idx
                        elif 'NOMBRE' in val and 'SECCION' in val:
                            col_map['nombre_seccion'] = cell.column - 1
                        elif 'IDENTIFICACION' in val or 'IDENTIFICACIÓN' in val:
                            col_map['codigo'] = cell.column - 1
                        elif val == 'FUENTE' or 'COD' in val and 'FTE' in val:
                            if 'cod_fuente' not in col_map:
                                col_map['cod_fuente'] = cell.column - 1
                        elif 'NOMBRE' in val and 'FUENTE' in val:
                            col_map['nombre_fuente'] = cell.column - 1
                        elif 'DESCRIPCION' in val or 'DESCRIPCIÓN' in val:
                            col_map['descripcion'] = cell.column - 1
                        elif 'APROPIACION' in val or 'APROPIACIÓN' in val:
                            col_map['apropiacion'] = cell.column - 1
                    if header_row:
                        break

                if not header_row or 'codigo' not in col_map:
                    messages.error(request, 'No se pudo detectar la estructura del archivo. Verifique que tenga columnas de IDENTIFICACIÓN PRESUPUESTAL y APROPIACIÓN.')
                    return redirect('importar_anexo2')

                # Borrar rubros existentes de esta vigencia antes de importar
                RubroGasto.objects.filter(vigencia=vigencia).delete()
                SeccionGasto.objects.filter(vigencia=vigencia).delete()
                FuenteFinanciacion.objects.filter(vigencia=vigencia).delete()

                count = 0
                orden = 0
                seccion_cache = {}
                fuente_cache = {}
                parent_stack = {}  # nivel -> último rubro título de ese nivel

                for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                    codigo_raw = str(row[col_map['codigo']] or '').strip() if 'codigo' in col_map and row[col_map['codigo']] else ''
                    descripcion = str(row[col_map.get('descripcion', col_map.get('codigo', 0))] or '').strip() if col_map.get('descripcion') is not None and len(row) > col_map.get('descripcion', 0) else ''

                    if not codigo_raw and not descripcion:
                        continue
                    if not codigo_raw:
                        continue

                    # Obtener sección
                    seccion_cod = str(row[col_map['seccion']] or '').strip() if 'seccion' in col_map and len(row) > col_map.get('seccion', 0) else ''
                    seccion_nombre = str(row[col_map['nombre_seccion']] or '').strip() if 'nombre_seccion' in col_map and len(row) > col_map.get('nombre_seccion', 0) else ''
                    seccion_obj = None
                    if seccion_cod:
                        if seccion_cod not in seccion_cache:
                            seccion_obj, _ = SeccionGasto.objects.get_or_create(
                                vigencia=vigencia, codigo=seccion_cod,
                                defaults={'nombre': seccion_nombre or seccion_cod}
                            )
                            seccion_cache[seccion_cod] = seccion_obj
                        else:
                            seccion_obj = seccion_cache[seccion_cod]

                    # Obtener fuente
                    cod_fuente = str(row[col_map['cod_fuente']] or '').strip() if 'cod_fuente' in col_map and len(row) > col_map.get('cod_fuente', 0) else ''
                    nombre_fuente = str(row[col_map['nombre_fuente']] or '').strip() if 'nombre_fuente' in col_map and len(row) > col_map.get('nombre_fuente', 0) else ''
                    fuente_obj = None
                    if cod_fuente:
                        if cod_fuente not in fuente_cache:
                            fuente_obj, _ = FuenteFinanciacion.objects.get_or_create(
                                vigencia=vigencia, codigo=cod_fuente,
                                defaults={'nombre': nombre_fuente or cod_fuente}
                            )
                            fuente_cache[cod_fuente] = fuente_obj
                        else:
                            fuente_obj = fuente_cache[cod_fuente]

                    # Apropiación
                    val_aprop = Decimal('0')
                    if 'apropiacion' in col_map and len(row) > col_map['apropiacion']:
                        val_aprop = _safe_decimal(row[col_map['apropiacion']])

                    # Determinar nivel jerárquico por el código
                    parts = codigo_raw.replace('-', '.').split('.')
                    nivel = len(parts) - 1
                    es_titulo = nivel < 3 and val_aprop == 0  # Los niveles altos sin valor son títulos

                    # Si tiene hijos potenciales (pocos segmentos), es título
                    if nivel <= 2:
                        es_titulo = True

                    # Determinar tipo de gasto
                    tipo_gasto = 'FUN'
                    desc_upper = descripcion.upper()
                    if 'INVERSIÓN' in desc_upper or 'INVERSION' in desc_upper:
                        tipo_gasto = 'INV'
                    elif 'DEUDA' in desc_upper:
                        tipo_gasto = 'DEU'

                    # Buscar parent
                    parent = None
                    if nivel > 0:
                        parent = parent_stack.get(nivel - 1)

                    orden += 1
                    rubro = RubroGasto.objects.create(
                        vigencia=vigencia,
                        codigo=codigo_raw,
                        descripcion=descripcion,
                        seccion=seccion_obj,
                        fuente=fuente_obj,
                        codigo_fuente=cod_fuente,
                        nombre_fuente=nombre_fuente,
                        tipo_gasto=tipo_gasto,
                        parent=parent,
                        valor_apropiacion=val_aprop if not es_titulo else 0,
                        es_titulo=es_titulo,
                        orden=orden,
                        nivel=nivel,
                    )

                    if es_titulo:
                        parent_stack[nivel] = rubro

                    count += 1

                # Recalcular títulos de abajo hacia arriba
                titulos = RubroGasto.objects.filter(vigencia=vigencia, es_titulo=True).order_by('-nivel')
                for titulo in titulos:
                    titulo.calcular_hijos()

                messages.success(request, f'{count} rubros de gasto importados desde Anexo 2')
                return redirect('rubros_gasto_list')
            except Exception as e:
                messages.error(request, f'Error al importar: {e}')
    else:
        form = ImportarGastosExcelForm()
    return render(request, 'gastos/importar_gastos.html', {
        'form': form,
        'tipo': 'Anexo 2 - Presupuesto de Gastos',
        'descripcion': 'Importe el archivo Excel del Anexo 2 (Presupuesto de Gastos Central). El sistema detectará automáticamente las columnas: Sección, Identificación Presupuestal, Fuente, Descripción y Apropiación.',
    })


# ─── IMPORTAR EJECUCIÓN DE GASTOS ─────────────────────────────────
@login_required
def importar_ejecucion(request):
    if request.method == 'POST':
        form = ImportarGastosExcelForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES['archivo']
            vigencia = _vigencia()
            try:
                wb = openpyxl.load_workbook(archivo, data_only=True)
                # Buscar la hoja de gastos
                sheet_name = None
                for name in wb.sheetnames:
                    if 'GASTO' in name.upper():
                        sheet_name = name
                        break
                if not sheet_name:
                    sheet_name = wb.sheetnames[0]
                ws = wb[sheet_name]

                # Detectar columnas por encabezados
                header_row = None
                col_map = {}
                for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=False), 1):
                    for cell in row:
                        val = str(cell.value or '').upper().strip()
                        if 'RUBRO' in val and 'PPTAL' in val:
                            col_map['codigo'] = cell.column - 1
                            header_row = row_idx
                        elif 'DESCRIPCI' in val and 'RUBRO' in val:
                            col_map['descripcion'] = cell.column - 1
                        elif 'PRESUPUESTO' in val and 'APROBADO' in val:
                            col_map['aprobado'] = cell.column - 1
                        elif val == 'ADICIONES':
                            col_map['adiciones'] = cell.column - 1
                        elif val == 'REDUCCIONES':
                            col_map['reducciones'] = cell.column - 1
                        elif 'TRASLADO' in val and 'CREDITO' in val and 'CONTRA' not in val:
                            col_map['traslado_credito'] = cell.column - 1
                        elif 'CONTRA' in val and ('CREDITO' in val or 'CRÉDITO' in val):
                            col_map['traslado_contra'] = cell.column - 1
                        elif 'APLAZAMIENTO' in val and 'DES' not in val:
                            col_map['aplazamientos'] = cell.column - 1
                        elif 'DESAPLAZAMIENTO' in val:
                            col_map['desaplazamientos'] = cell.column - 1
                        elif val == 'CDP' or 'CDP' in val and 'SALDO' not in val and 'COMPROMETER' not in val:
                            if 'cdp' not in col_map:
                                col_map['cdp'] = cell.column - 1
                        elif 'COMPROMISO' in val:
                            col_map['compromisos'] = cell.column - 1
                        elif 'ORDENADO' in val or 'OBLIGACION' in val:
                            col_map['ordenado'] = cell.column - 1
                        elif 'PAGADO' in val or 'PAGO' in val:
                            col_map['pagado'] = cell.column - 1
                        elif 'FUENTE' in val and 'PPTAL' in val and 'DESCRIPCI' not in val:
                            col_map['fuente'] = cell.column - 1
                        elif 'DESCRIPCI' in val and 'FUENTE' in val:
                            col_map['nombre_fuente'] = cell.column - 1
                        elif 'TIPO' in val and 'GASTO' in val:
                            col_map['tipo_gasto'] = cell.column - 1
                    if header_row:
                        break

                if not header_row or 'codigo' not in col_map:
                    messages.error(request, 'No se pudo detectar la estructura del archivo de ejecución.')
                    return redirect('importar_ejecucion')

                count_new = 0
                count_updated = 0

                for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                    codigo = str(row[col_map['codigo']] or '').strip() if len(row) > col_map['codigo'] else ''
                    if not codigo:
                        continue

                    descripcion = ''
                    if 'descripcion' in col_map and len(row) > col_map['descripcion']:
                        descripcion = str(row[col_map['descripcion']] or '').strip()

                    # Buscar rubro existente o crear nuevo
                    rubro = RubroGasto.objects.filter(vigencia=vigencia, codigo=codigo).first()
                    if not rubro:
                        parts = codigo.replace('-', '.').split('.')
                        nivel = len(parts) - 1
                        rubro = RubroGasto.objects.create(
                            vigencia=vigencia,
                            codigo=codigo,
                            descripcion=descripcion,
                            nivel=nivel,
                            es_titulo=False,
                            orden=RubroGasto.objects.filter(vigencia=vigencia).count() + 1,
                        )
                        count_new += 1

                    # Crear o actualizar ejecución
                    ej, created = EjecucionGasto.objects.get_or_create(rubro=rubro)
                    ej.presupuesto_aprobado = _safe_decimal(row[col_map['aprobado']]) if 'aprobado' in col_map and len(row) > col_map['aprobado'] else ej.presupuesto_aprobado
                    ej.adiciones = _safe_decimal(row[col_map['adiciones']]) if 'adiciones' in col_map and len(row) > col_map['adiciones'] else ej.adiciones
                    ej.reducciones = _safe_decimal(row[col_map['reducciones']]) if 'reducciones' in col_map and len(row) > col_map['reducciones'] else ej.reducciones
                    ej.traslado_credito = _safe_decimal(row[col_map['traslado_credito']]) if 'traslado_credito' in col_map and len(row) > col_map['traslado_credito'] else ej.traslado_credito
                    ej.traslado_contra_credito = _safe_decimal(row[col_map['traslado_contra']]) if 'traslado_contra' in col_map and len(row) > col_map['traslado_contra'] else ej.traslado_contra_credito
                    ej.aplazamientos = _safe_decimal(row[col_map['aplazamientos']]) if 'aplazamientos' in col_map and len(row) > col_map['aplazamientos'] else ej.aplazamientos
                    ej.desaplazamientos = _safe_decimal(row[col_map['desaplazamientos']]) if 'desaplazamientos' in col_map and len(row) > col_map['desaplazamientos'] else ej.desaplazamientos
                    ej.cdp = _safe_decimal(row[col_map['cdp']]) if 'cdp' in col_map and len(row) > col_map['cdp'] else ej.cdp
                    ej.compromisos = _safe_decimal(row[col_map['compromisos']]) if 'compromisos' in col_map and len(row) > col_map['compromisos'] else ej.compromisos
                    ej.ordenado = _safe_decimal(row[col_map['ordenado']]) if 'ordenado' in col_map and len(row) > col_map['ordenado'] else ej.ordenado
                    ej.pagado = _safe_decimal(row[col_map['pagado']]) if 'pagado' in col_map and len(row) > col_map['pagado'] else ej.pagado
                    ej.save()
                    count_updated += 1

                messages.success(request, f'Ejecución importada: {count_updated} registros actualizados, {count_new} rubros nuevos creados')
                return redirect('ejecucion_gastos')
            except Exception as e:
                messages.error(request, f'Error al importar ejecución: {e}')
    else:
        form = ImportarGastosExcelForm()
    return render(request, 'gastos/importar_gastos.html', {
        'form': form,
        'tipo': 'Ejecución de Gastos',
        'descripcion': 'Importe el archivo Excel de Ejecución de Gastos. El sistema detectará automáticamente las columnas: Rubro, Presupuesto Aprobado, Adiciones, Reducciones, CDP, Compromisos, Ordenado y Pagado.',
    })


# ─── EJECUCIÓN DE GASTOS ─────────────────────────────────────────
@login_required
def ejecucion_gastos(request):
    vigencia = _vigencia()
    rubros_con_ejecucion = RubroGasto.objects.filter(
        vigencia=vigencia, es_titulo=False, ejecucion__isnull=False
    ).select_related('ejecucion', 'seccion')

    totales = EjecucionGasto.objects.filter(rubro__vigencia=vigencia).aggregate(
        total_aprobado=Sum('presupuesto_aprobado'),
        total_adiciones=Sum('adiciones'),
        total_reducciones=Sum('reducciones'),
        total_cdp=Sum('cdp'),
        total_compromisos=Sum('compromisos'),
        total_ordenado=Sum('ordenado'),
        total_pagado=Sum('pagado'),
    )

    total_aprobado = totales['total_aprobado'] or 0
    total_compromisos = totales['total_compromisos'] or 0
    pct_ejecucion = (Decimal(str(total_compromisos)) / Decimal(str(total_aprobado)) * 100).quantize(Decimal('0.01')) if total_aprobado else Decimal('0')

    return render(request, 'gastos/ejecucion_gastos.html', {
        'rubros': rubros_con_ejecucion,
        'vigencia': vigencia,
        'totales': totales,
        'pct_ejecucion': pct_ejecucion,
    })


# ─── EDITAR EJECUCIÓN INDIVIDUAL ──────────────────────────────────
@login_required
def ejecucion_editar(request, pk):
    rubro = get_object_or_404(RubroGasto, pk=pk)
    ejecucion, created = EjecucionGasto.objects.get_or_create(rubro=rubro)
    if request.method == 'POST':
        form = EjecucionGastoForm(request.POST, instance=ejecucion)
        if form.is_valid():
            form.save()
            messages.success(request, f'Ejecución de {rubro.codigo} actualizada')
            return redirect('ejecucion_gastos')
    else:
        form = EjecucionGastoForm(instance=ejecucion)
    return render(request, 'gastos/ejecucion_form.html', {
        'form': form, 'rubro': rubro,
    })


# ─── LIMPIAR / BORRAR DATOS ───────────────────────────────────────
@login_required
def limpiar_gastos(request):
    vigencia = _vigencia()
    rubros_count = RubroGasto.objects.filter(vigencia=vigencia).count()
    ejecucion_count = EjecucionGasto.objects.filter(rubro__vigencia=vigencia).count()
    secciones_count = SeccionGasto.objects.filter(vigencia=vigencia).count()
    fuentes_count = FuenteFinanciacion.objects.filter(vigencia=vigencia).count()

    if request.method == 'POST':
        accion = request.POST.get('accion', '')

        if accion == 'borrar_rubros':
            n, _ = RubroGasto.objects.filter(vigencia=vigencia).delete()
            messages.success(request, f'Se eliminaron {n} registros de rubros de gasto (y sus ejecuciones asociadas)')

        elif accion == 'borrar_ejecucion':
            n, _ = EjecucionGasto.objects.filter(rubro__vigencia=vigencia).delete()
            messages.success(request, f'Se eliminaron {n} registros de ejecución')

        elif accion == 'borrar_secciones':
            n, _ = SeccionGasto.objects.filter(vigencia=vigencia).delete()
            messages.success(request, f'Se eliminaron {n} secciones')

        elif accion == 'borrar_fuentes':
            n, _ = FuenteFinanciacion.objects.filter(vigencia=vigencia).delete()
            messages.success(request, f'Se eliminaron {n} fuentes')

        elif accion == 'borrar_todo':
            EjecucionGasto.objects.filter(rubro__vigencia=vigencia).delete()
            RubroGasto.objects.filter(vigencia=vigencia).delete()
            SeccionGasto.objects.filter(vigencia=vigencia).delete()
            FuenteFinanciacion.objects.filter(vigencia=vigencia).delete()
            messages.success(request, f'Se eliminaron TODOS los datos de gastos de la vigencia {vigencia}')

        return redirect('limpiar_gastos')

    return render(request, 'gastos/limpiar_gastos.html', {
        'vigencia': vigencia,
        'rubros_count': rubros_count,
        'ejecucion_count': ejecucion_count,
        'secciones_count': secciones_count,
        'fuentes_count': fuentes_count,
    })


# ─── REPORTE ANEXO 2 ─────────────────────────────────────────────
@login_required
def reporte_gastos(request):
    vigencia = _vigencia()
    params = ParametrosSistema.objects.filter(vigencia=vigencia).first()
    rubros = RubroGasto.objects.filter(vigencia=vigencia)
    total = rubros.filter(nivel=0, es_titulo=True).aggregate(t=Sum('valor_apropiacion'))['t']
    if not total:
        total = rubros.filter(es_titulo=False).aggregate(t=Sum('valor_apropiacion'))['t'] or 0

    # Totales por tipo de gasto
    total_funcionamiento = rubros.filter(tipo_gasto='FUN', es_titulo=False).aggregate(t=Sum('valor_apropiacion'))['t'] or 0
    total_inversion = rubros.filter(tipo_gasto='INV', es_titulo=False).aggregate(t=Sum('valor_apropiacion'))['t'] or 0
    total_deuda = rubros.filter(tipo_gasto='DEU', es_titulo=False).aggregate(t=Sum('valor_apropiacion'))['t'] or 0

    return render(request, 'gastos/reporte_gastos.html', {
        'rubros': rubros, 'vigencia': vigencia, 'params': params, 'total': total,
        'total_funcionamiento': total_funcionamiento,
        'total_inversion': total_inversion,
        'total_deuda': total_deuda,
    })


# ─── EXPORTAR EXCEL ──────────────────────────────────────────────
@login_required
def exportar_gastos_excel(request):
    vigencia = _vigencia()
    rubros = RubroGasto.objects.filter(vigencia=vigencia)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Anexo 2 Detalle Gastos'
    ws.append(['MUNICIPIO DE PUERTO LÓPEZ'])
    ws.append(['PRESUPUESTO DE GASTOS - ADMINISTRACIÓN CENTRAL'])
    ws.append([f'VIGENCIA FISCAL {vigencia}'])
    ws.append([])
    ws.append(['SECCIÓN', 'IDENTIFICACIÓN', 'COD FUENTE', 'FUENTE', 'DESCRIPCIÓN', 'APROPIACIÓN', 'OBSERVACIONES'])

    for rubro in rubros:
        indent = '  ' * rubro.nivel
        ws.append([
            rubro.seccion.codigo if rubro.seccion else '',
            rubro.codigo,
            rubro.codigo_fuente,
            rubro.nombre_fuente,
            indent + rubro.descripcion,
            float(rubro.valor_apropiacion),
            rubro.observaciones,
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=Anexo2_Gastos_{vigencia}.xlsx'
    wb.save(response)
    return response


# ─── EXPORTAR EJECUCIÓN EXCEL ────────────────────────────────────
@login_required
def exportar_ejecucion_excel(request):
    vigencia = _vigencia()
    rubros = RubroGasto.objects.filter(
        vigencia=vigencia, es_titulo=False, ejecucion__isnull=False
    ).select_related('ejecucion')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Ejecución Gastos'
    ws.append(['MUNICIPIO DE PUERTO LÓPEZ'])
    ws.append([f'EJECUCIÓN PRESUPUESTAL DE GASTOS - VIGENCIA {vigencia}'])
    ws.append([])
    ws.append([
        'CÓDIGO', 'DESCRIPCIÓN', 'PPTO APROBADO', 'ADICIONES', 'REDUCCIONES',
        'TRASLADO CRÉDITO', 'CONTRA CRÉDITO', 'APLAZAMIENTOS', 'DESAPLAZAMIENTOS',
        'APROPIACIÓN DEFINITIVA', 'CDP', 'COMPROMISOS', 'ORDENADO', 'PAGADO',
        'SALDO APROPIACIÓN', '% EJECUCIÓN'
    ])

    for r in rubros:
        ej = r.ejecucion
        ws.append([
            r.codigo, r.descripcion,
            float(ej.presupuesto_aprobado), float(ej.adiciones), float(ej.reducciones),
            float(ej.traslado_credito), float(ej.traslado_contra_credito),
            float(ej.aplazamientos), float(ej.desaplazamientos),
            float(ej.apropiacion_definitiva),
            float(ej.cdp), float(ej.compromisos), float(ej.ordenado), float(ej.pagado),
            float(ej.saldo_apropiacion), float(ej.porcentaje_ejecucion),
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=Ejecucion_Gastos_{vigencia}.xlsx'
    wb.save(response)
    return response


# ─── CIFRAS HISTÓRICAS GASTOS ────────────────────────────────────
def _calcular_tcpa(valores_por_anio):
    """Calcula la Tasa Compuesta Promedio Anual."""
    if len(valores_por_anio) < 2:
        return Decimal('0')
    anios = sorted(valores_por_anio.keys())
    v0 = valores_por_anio[anios[0]]
    vn = valores_por_anio[anios[-1]]
    n = len(anios) - 1
    if v0 <= 0 or vn <= 0:
        return Decimal('0')
    ratio = float(vn) / float(v0)
    tcpa = ratio ** (1.0 / n) - 1.0
    return Decimal(str(round(tcpa, 6)))


@login_required
def cifras_historicas_gastos(request):
    vigencia = _vigencia()
    cifras = CifraHistoricaGasto.objects.filter(vigencia_calculo=vigencia)
    anios = sorted(set(cifras.values_list('anio', flat=True)))

    # Totales por año
    totales_por_anio = {}
    totales_fun_por_anio = {}
    totales_inv_por_anio = {}
    totales_deu_por_anio = {}
    pct_pagos_por_anio = {}
    for anio in anios:
        cifras_anio = cifras.filter(anio=anio)
        agg = cifras_anio.aggregate(
            t_aprop=Sum('valor_apropiacion'),
            t_pagos=Sum('valor_pagos'),
        )
        totales_por_anio[anio] = agg['t_aprop'] or Decimal('0')
        total_pagos = agg['t_pagos'] or Decimal('0')
        if totales_por_anio[anio] > 0:
            pct_pagos_por_anio[anio] = (total_pagos / totales_por_anio[anio] * 100).quantize(Decimal('0.01'))
        else:
            pct_pagos_por_anio[anio] = Decimal('0')
        totales_fun_por_anio[anio] = cifras_anio.filter(tipo_gasto='FUN').aggregate(t=Sum('valor_apropiacion'))['t'] or Decimal('0')
        totales_inv_por_anio[anio] = cifras_anio.filter(tipo_gasto='INV').aggregate(t=Sum('valor_apropiacion'))['t'] or Decimal('0')
        totales_deu_por_anio[anio] = cifras_anio.filter(tipo_gasto='DEU').aggregate(t=Sum('valor_apropiacion'))['t'] or Decimal('0')

    tcpa = _calcular_tcpa(totales_por_anio)

    # % promedio de pagos sobre apropiación
    if pct_pagos_por_anio:
        pct_promedio = sum(pct_pagos_por_anio.values()) / len(pct_pagos_por_anio)
    else:
        pct_promedio = Decimal('0')

    # Tabla Concejo/Personería
    params = ParametrosSistema.objects.filter(vigencia=vigencia).first()
    tabla_cp = TablaConcejoPersoneria.objects.all()
    tabla_actual = None
    honorarios_concejo = Decimal('0')
    limite_concejo = Decimal('0')
    limite_personeria = Decimal('0')
    if params:
        tabla_actual = TablaConcejoPersoneria.objects.filter(categoria=params.categoria_municipio).first()
        if tabla_actual:
            honorarios_concejo = tabla_actual.calcular_honorarios_concejo(params.valor_smlmv)
            # Necesitamos ICLD para calcular límites
            from ingresos.models import CifraHistoricaIngreso
            cifras_ing = CifraHistoricaIngreso.objects.filter(vigencia_calculo=vigencia)
            ultimo_anio_ing = cifras_ing.values_list('anio', flat=True).order_by('-anio').first()
            if ultimo_anio_ing:
                icld = cifras_ing.filter(anio=ultimo_anio_ing, es_icld=True).aggregate(t=Sum('valor_recaudo'))['t'] or Decimal('0')
                sgp_libre = cifras_ing.filter(anio=ultimo_anio_ing, es_sgp_libre=True).aggregate(t=Sum('valor_recaudo'))['t'] or Decimal('0')
                icld_total = icld + sgp_libre
                limite_concejo = tabla_actual.calcular_limite_concejo(icld_total)
                limite_personeria = tabla_actual.calcular_limite_personeria(icld_total)

    form = CifraHistoricaGastoForm(initial={'vigencia_calculo': vigencia})

    return render(request, 'gastos/cifras_historicas_gastos.html', {
        'cifras': cifras,
        'anios': anios,
        'totales_por_anio': totales_por_anio,
        'totales_fun_por_anio': totales_fun_por_anio,
        'totales_inv_por_anio': totales_inv_por_anio,
        'totales_deu_por_anio': totales_deu_por_anio,
        'pct_pagos_por_anio': pct_pagos_por_anio,
        'pct_promedio': pct_promedio,
        'tcpa': tcpa,
        'tabla_cp': tabla_cp,
        'tabla_actual': tabla_actual,
        'honorarios_concejo': honorarios_concejo,
        'limite_concejo': limite_concejo,
        'limite_personeria': limite_personeria,
        'vigencia': vigencia,
        'params': params,
        'form': form,
    })


@login_required
def cifra_historica_gasto_guardar(request):
    if request.method == 'POST':
        pk = request.POST.get('pk')
        instance = get_object_or_404(CifraHistoricaGasto, pk=pk) if pk else None
        form = CifraHistoricaGastoForm(request.POST, instance=instance)
        if form.is_valid():
            form.save()
            messages.success(request, 'Cifra histórica guardada')
    return redirect('cifras_historicas_gastos')


@login_required
def cifra_historica_gasto_eliminar(request, pk):
    get_object_or_404(CifraHistoricaGasto, pk=pk).delete()
    messages.success(request, 'Cifra histórica eliminada')
    return redirect('cifras_historicas_gastos')


@login_required
def importar_cifras_historicas_gastos(request):
    """Importa cifras históricas de gastos desde Excel."""
    if request.method == 'POST':
        form = ImportarGastosExcelForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES['archivo']
            vigencia = _vigencia()
            try:
                wb = openpyxl.load_workbook(archivo, data_only=True)
                ws = wb.active
                header = [str(c.value or '').strip() for c in ws[1]]
                anio_cols = {}
                for idx, h in enumerate(header):
                    try:
                        anio = int(h)
                        if 2020 <= anio <= 2030:
                            anio_cols[anio] = idx
                    except ValueError:
                        pass

                count = 0
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or not row[0]:
                        continue
                    codigo = str(row[0]).strip()
                    descripcion = str(row[1] or '').strip() if len(row) > 1 else ''
                    tipo_gasto = 'FUN'
                    for idx, h in enumerate(header):
                        if 'TIPO' in h.upper() and len(row) > idx:
                            val = str(row[idx] or '').upper().strip()
                            if val in ('FUN', 'INV', 'DEU'):
                                tipo_gasto = val
                            elif 'INV' in val:
                                tipo_gasto = 'INV'
                            elif 'DEU' in val:
                                tipo_gasto = 'DEU'

                    for anio, col_idx in anio_cols.items():
                        if col_idx < len(row) and row[col_idx]:
                            try:
                                valor = Decimal(str(row[col_idx]).replace(',', '').replace('$', '').strip())
                            except Exception:
                                continue
                            CifraHistoricaGasto.objects.update_or_create(
                                vigencia_calculo=vigencia, anio=anio, codigo_rubro=codigo,
                                defaults={
                                    'descripcion': descripcion,
                                    'valor_apropiacion': valor,
                                    'tipo_gasto': tipo_gasto,
                                }
                            )
                            count += 1
                messages.success(request, f'{count} cifras históricas de gastos importadas')
                return redirect('cifras_historicas_gastos')
            except Exception as e:
                messages.error(request, f'Error al importar: {e}')
    else:
        form = ImportarGastosExcelForm()
    return render(request, 'gastos/importar_gastos.html', {
        'form': form,
        'tipo': 'Cifras Históricas Gastos CUIPO',
        'descripcion': 'Importe Excel con columnas: Código, Descripción, Tipo (FUN/INV/DEU), 2022, 2023, 2024, 2025',
    })


@login_required
def calcular_tcpa_gastos(request):
    vigencia = _vigencia()
    cifras = CifraHistoricaGasto.objects.filter(vigencia_calculo=vigencia)
    anios = sorted(set(cifras.values_list('anio', flat=True)))
    totales_por_anio = {}
    for anio in anios:
        totales_por_anio[anio] = cifras.filter(anio=anio).aggregate(t=Sum('valor_apropiacion'))['t'] or Decimal('0')

    tcpa = _calcular_tcpa(totales_por_anio)
    params = ParametrosSistema.objects.filter(vigencia=vigencia).first()
    if params:
        params.tcpa_gastos = tcpa
        params.save(update_fields=['tcpa_gastos'])

        # También calcular y guardar % promedio pagos
        pct_pagos = []
        for anio in anios:
            agg = cifras.filter(anio=anio).aggregate(t_aprop=Sum('valor_apropiacion'), t_pagos=Sum('valor_pagos'))
            if agg['t_aprop'] and agg['t_aprop'] > 0:
                pct_pagos.append(agg['t_pagos'] / agg['t_aprop'] * 100)
        if pct_pagos:
            params.pct_promedio_pagos = sum(pct_pagos) / len(pct_pagos)
            params.save(update_fields=['pct_promedio_pagos'])

        messages.success(request, f'TCPA Gastos calculada: {tcpa * 100:.2f}%')
    else:
        messages.error(request, 'No hay parámetros configurados')
    return redirect('cifras_historicas_gastos')


# ─── DEUDA PÚBLICA - CONTRATOS ───────────────────────────────────
@login_required
def deuda_contratos_list(request):
    vigencia = _vigencia()
    contratos = ContratoCredito.objects.filter(vigencia=vigencia).prefetch_related('pagares')
    form = ContratoCreditoForm(initial={'vigencia': vigencia})
    return render(request, 'gastos/deuda_contratos.html', {
        'contratos': contratos, 'form': form, 'vigencia': vigencia,
    })


@login_required
def deuda_contrato_guardar(request):
    if request.method == 'POST':
        pk = request.POST.get('pk')
        instance = get_object_or_404(ContratoCredito, pk=pk) if pk else None
        form = ContratoCreditoForm(request.POST, instance=instance)
        if form.is_valid():
            form.save()
            messages.success(request, 'Contrato de crédito guardado')
    return redirect('deuda_contratos_list')


@login_required
def deuda_contrato_eliminar(request, pk):
    get_object_or_404(ContratoCredito, pk=pk).delete()
    messages.success(request, 'Contrato eliminado')
    return redirect('deuda_contratos_list')


# ─── DEUDA PÚBLICA - PAGARÉS ────────────────────────────────────
@login_required
def deuda_pagares(request, contrato_pk):
    contrato = get_object_or_404(ContratoCredito, pk=contrato_pk)
    pagares = contrato.pagares.all()
    form = PagareCreditoForm(initial={'contrato': contrato})
    form.fields['contrato'].queryset = ContratoCredito.objects.filter(pk=contrato_pk)
    return render(request, 'gastos/deuda_pagares.html', {
        'contrato': contrato, 'pagares': pagares, 'form': form,
    })


@login_required
def deuda_pagare_guardar(request):
    if request.method == 'POST':
        pk = request.POST.get('pk')
        instance = get_object_or_404(PagareCredito, pk=pk) if pk else None
        form = PagareCreditoForm(request.POST, instance=instance)
        if form.is_valid():
            pagare = form.save()
            messages.success(request, f'Pagaré {pagare.numero_pagare} guardado')
            return redirect('deuda_pagares', contrato_pk=pagare.contrato.pk)
    return redirect('deuda_contratos_list')


@login_required
def deuda_pagare_eliminar(request, pk):
    pagare = get_object_or_404(PagareCredito, pk=pk)
    contrato_pk = pagare.contrato.pk
    pagare.delete()
    messages.success(request, 'Pagaré eliminado')
    return redirect('deuda_pagares', contrato_pk=contrato_pk)


# ─── DEUDA PÚBLICA - AMORTIZACIÓN ───────────────────────────────
@login_required
def deuda_amortizacion(request, pagare_pk):
    pagare = get_object_or_404(PagareCredito, pk=pagare_pk)
    amortizaciones = pagare.amortizaciones.all()

    if request.method == 'POST':
        # Guardar todas las filas de amortización del formulario
        vigencias = request.POST.getlist('vigencia_pago')
        capitales = request.POST.getlist('capital_principal')
        intereses_list = request.POST.getlist('intereses')
        tcr_list = request.POST.getlist('intereses_tcr')

        pagare.amortizaciones.all().delete()
        for i in range(len(vigencias)):
            if vigencias[i]:
                AmortizacionPagare.objects.create(
                    pagare=pagare,
                    vigencia_pago=int(vigencias[i]),
                    capital_principal=_safe_decimal(capitales[i]) if i < len(capitales) else 0,
                    intereses=_safe_decimal(intereses_list[i]) if i < len(intereses_list) else 0,
                    intereses_tcr=_safe_decimal(tcr_list[i]) if i < len(tcr_list) else 0,
                )
        messages.success(request, f'Tabla de amortización del pagaré {pagare.numero_pagare} guardada')
        return redirect('deuda_amortizacion', pagare_pk=pagare.pk)

    # Generar vigencias vacías si no existen (2026-2036)
    vigencias_default = list(range(2026, 2037))
    amort_dict = {a.vigencia_pago: a for a in amortizaciones}
    tabla = []
    for v in vigencias_default:
        if v in amort_dict:
            a = amort_dict[v]
            tabla.append({'vigencia': v, 'capital': a.capital_principal, 'intereses': a.intereses, 'tcr': a.intereses_tcr, 'total': a.total})
        else:
            tabla.append({'vigencia': v, 'capital': Decimal('0'), 'intereses': Decimal('0'), 'tcr': Decimal('0'), 'total': Decimal('0')})

    totales = {
        'capital': sum(r['capital'] for r in tabla),
        'intereses': sum(r['intereses'] for r in tabla),
        'tcr': sum(r['tcr'] for r in tabla),
        'total': sum(r['total'] for r in tabla),
    }

    return render(request, 'gastos/deuda_amortizacion.html', {
        'pagare': pagare, 'contrato': pagare.contrato,
        'tabla': tabla, 'totales': totales,
    })


# ─── DEUDA PÚBLICA - RESUMEN CONSOLIDADO ─────────────────────────
@login_required
def deuda_resumen(request):
    vigencia = _vigencia()
    contratos = ContratoCredito.objects.filter(vigencia=vigencia).prefetch_related('pagares__amortizaciones')

    vigencias_rango = list(range(2026, 2037))

    # Consolidar por contrato (banco)
    resumen_contratos = []
    gran_total = {v: {'capital': Decimal('0'), 'intereses': Decimal('0'), 'tcr': Decimal('0'), 'total': Decimal('0')} for v in vigencias_rango}

    for contrato in contratos:
        consolidado = {v: {'capital': Decimal('0'), 'intereses': Decimal('0'), 'tcr': Decimal('0'), 'total': Decimal('0')} for v in vigencias_rango}
        for pagare in contrato.pagares.all():
            for amort in pagare.amortizaciones.all():
                if amort.vigencia_pago in consolidado:
                    consolidado[amort.vigencia_pago]['capital'] += amort.capital_principal
                    consolidado[amort.vigencia_pago]['intereses'] += amort.intereses
                    consolidado[amort.vigencia_pago]['tcr'] += amort.intereses_tcr
                    consolidado[amort.vigencia_pago]['total'] += amort.total
                    gran_total[amort.vigencia_pago]['capital'] += amort.capital_principal
                    gran_total[amort.vigencia_pago]['intereses'] += amort.intereses
                    gran_total[amort.vigencia_pago]['tcr'] += amort.intereses_tcr
                    gran_total[amort.vigencia_pago]['total'] += amort.total

        total_contrato = {
            'capital': sum(v['capital'] for v in consolidado.values()),
            'intereses': sum(v['intereses'] for v in consolidado.values()),
            'tcr': sum(v['tcr'] for v in consolidado.values()),
            'total': sum(v['total'] for v in consolidado.values()),
        }

        resumen_contratos.append({
            'contrato': contrato,
            'consolidado': consolidado,
            'total': total_contrato,
        })

    # Total para presupuesto (solo vigencia actual)
    total_presupuesto = gran_total.get(vigencia, {'capital': Decimal('0'), 'intereses': Decimal('0'), 'tcr': Decimal('0'), 'total': Decimal('0')})

    return render(request, 'gastos/deuda_resumen.html', {
        'vigencia': vigencia, 'vigencias_rango': vigencias_rango,
        'resumen_contratos': resumen_contratos,
        'gran_total': gran_total,
        'total_presupuesto': total_presupuesto,
    })


# ─── COSTO DE PERSONAL ──────────────────────────────────────────
@login_required
def costo_personal_list(request):
    vigencia = _vigencia()
    personal = CostoPersonal.objects.filter(vigencia=vigencia).select_related('seccion')
    personal_activo = personal.filter(es_pensionado=False)
    pensionados = personal.filter(es_pensionado=True)

    total_salarios = sum(p.costo_salarial_anual for p in personal_activo)
    total_prestaciones = sum(p.costo_prestaciones for p in personal_activo)
    total_aportes = sum(p.costo_aportes for p in personal_activo)
    total_general = sum(p.costo_total_anual for p in personal_activo)
    total_pensionados = sum(p.costo_total_anual for p in pensionados)

    # Agrupar por sección
    secciones_resumen = {}
    for p in personal_activo:
        sec_nombre = p.seccion.nombre if p.seccion else 'Sin Sección'
        if sec_nombre not in secciones_resumen:
            secciones_resumen[sec_nombre] = Decimal('0')
        secciones_resumen[sec_nombre] += p.costo_total_anual

    form = CostoPersonalForm(initial={'vigencia': vigencia})
    form.fields['seccion'].queryset = SeccionGasto.objects.filter(vigencia=vigencia)

    return render(request, 'gastos/costo_personal.html', {
        'personal': personal_activo, 'pensionados': pensionados,
        'form': form, 'vigencia': vigencia,
        'total_salarios': total_salarios, 'total_prestaciones': total_prestaciones,
        'total_aportes': total_aportes, 'total_general': total_general,
        'total_pensionados': total_pensionados,
        'secciones_resumen': secciones_resumen,
    })


@login_required
def costo_personal_guardar(request):
    if request.method == 'POST':
        pk = request.POST.get('pk')
        instance = get_object_or_404(CostoPersonal, pk=pk) if pk else None
        form = CostoPersonalForm(request.POST, instance=instance)
        if form.is_valid():
            form.save()
            messages.success(request, 'Costo de personal guardado')
    return redirect('costo_personal_list')


@login_required
def costo_personal_eliminar(request, pk):
    get_object_or_404(CostoPersonal, pk=pk).delete()
    messages.success(request, 'Registro eliminado')
    return redirect('costo_personal_list')


# ─── VIGENCIAS FUTURAS ──────────────────────────────────────────
@login_required
def vigencias_futuras_list(request):
    from core.forms import VigenciaFuturaForm
    vigencia = _vigencia()
    vf = VigenciaFutura.objects.filter(vigencia=vigencia)
    vf_aprobadas = vf.filter(estado='APR')
    vf_ejecucion = vf.filter(estado='EJE')

    # Agrupar por fuente
    por_fuente = {}
    for item in vf:
        key = item.nombre_fuente or item.codigo_fuente
        if key not in por_fuente:
            por_fuente[key] = {'aprobadas': Decimal('0'), 'ejecucion': Decimal('0')}
        if item.estado == 'APR':
            por_fuente[key]['aprobadas'] += item.valor
        else:
            por_fuente[key]['ejecucion'] += item.valor

    total_aprobadas = vf_aprobadas.aggregate(t=Sum('valor'))['t'] or 0
    total_ejecucion = vf_ejecucion.aggregate(t=Sum('valor'))['t'] or 0

    form = VigenciaFuturaForm(initial={'vigencia': vigencia})

    return render(request, 'gastos/vigencias_futuras.html', {
        'vf': vf, 'form': form, 'vigencia': vigencia,
        'vf_aprobadas': vf_aprobadas, 'vf_ejecucion': vf_ejecucion,
        'por_fuente': por_fuente,
        'total_aprobadas': total_aprobadas, 'total_ejecucion': total_ejecucion,
    })


@login_required
def vigencia_futura_guardar(request):
    from core.forms import VigenciaFuturaForm
    if request.method == 'POST':
        pk = request.POST.get('pk')
        instance = get_object_or_404(VigenciaFutura, pk=pk) if pk else None
        form = VigenciaFuturaForm(request.POST, instance=instance)
        if form.is_valid():
            form.save()
            messages.success(request, 'Vigencia futura guardada')
    return redirect('vigencias_futuras_list')


@login_required
def vigencia_futura_eliminar(request, pk):
    get_object_or_404(VigenciaFutura, pk=pk).delete()
    messages.success(request, 'Registro eliminado')
    return redirect('vigencias_futuras_list')


# ─── REPORTE TECHOS DE INVERSIÓN ─────────────────────────────────
@login_required
def reporte_techos_inversion(request):
    vigencia = _vigencia()
    params = ParametrosSistema.objects.filter(vigencia=vigencia).first()

    # Rubros de inversión agrupados por fuente
    rubros_inv = RubroGasto.objects.filter(vigencia=vigencia, tipo_gasto='INV', es_titulo=False)
    por_fuente = {}
    for r in rubros_inv:
        key = r.nombre_fuente or r.codigo_fuente or 'Sin Fuente'
        if key not in por_fuente:
            por_fuente[key] = Decimal('0')
        por_fuente[key] += r.valor_apropiacion
    total_inversion = sum(por_fuente.values())

    # ICLD para cálculos específicos
    from ingresos.models import RubroIngreso, CifraHistoricaIngreso
    cifras_ing = CifraHistoricaIngreso.objects.filter(vigencia_calculo=vigencia)
    ultimo_anio = cifras_ing.values_list('anio', flat=True).order_by('-anio').first()
    icld_propios = Decimal('0')
    if ultimo_anio:
        icld_propios = cifras_ing.filter(anio=ultimo_anio, es_icld=True).aggregate(t=Sum('valor_recaudo'))['t'] or Decimal('0')

    # Cálculos FAEP
    # Predial total (buscar en rubros de ingreso)
    total_predial = RubroIngreso.objects.filter(
        vigencia=vigencia, es_titulo=False,
        metodo_calculo__in=['PUVA', 'PUAN', 'PRVA', 'PRAN']
    ).aggregate(t=Sum('valor_apropiacion'))['t'] or Decimal('0')

    faep_predial = total_predial * Decimal('0.10')  # 10% predial
    vivienda_predial = total_predial * Decimal('0.10')  # 10% vivienda predial
    medio_ambiente = icld_propios * Decimal('0.01')  # 1% ICLD propios

    # Vigencias futuras en ejecución
    vf_ejecucion = VigenciaFutura.objects.filter(vigencia=vigencia, estado='EJE')
    vf_aprobadas = VigenciaFutura.objects.filter(vigencia=vigencia, estado='APR')

    # Indicador Ley 617
    total_funcionamiento = RubroGasto.objects.filter(
        vigencia=vigencia, tipo_gasto='FUN', es_titulo=False
    ).aggregate(t=Sum('valor_apropiacion'))['t'] or Decimal('0')

    sgp_libre = Decimal('0')
    if ultimo_anio:
        sgp_libre = cifras_ing.filter(anio=ultimo_anio, es_sgp_libre=True).aggregate(t=Sum('valor_recaudo'))['t'] or Decimal('0')
    icld_total_617 = icld_propios + sgp_libre
    indicador_617 = Decimal('0')
    if icld_total_617 > 0:
        indicador_617 = (total_funcionamiento / icld_total_617 * 100).quantize(Decimal('0.01'))

    return render(request, 'gastos/reporte_techos.html', {
        'vigencia': vigencia, 'params': params,
        'por_fuente': por_fuente, 'total_inversion': total_inversion,
        'icld_propios': icld_propios,
        'total_predial': total_predial,
        'faep_predial': faep_predial,
        'vivienda_predial': vivienda_predial,
        'medio_ambiente': medio_ambiente,
        'vf_ejecucion': vf_ejecucion, 'vf_aprobadas': vf_aprobadas,
        'total_funcionamiento': total_funcionamiento,
        'icld_total_617': icld_total_617,
        'indicador_617': indicador_617,
    })
