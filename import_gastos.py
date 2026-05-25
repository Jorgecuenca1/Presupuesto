import os, sys, django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'presupuesto_project.settings')
django.setup()

from decimal import Decimal, InvalidOperation
import openpyxl
from django.db.models import Sum
from gastos.models import RubroGasto, SeccionGasto, FuenteFinanciacion, EjecucionGasto
from core.models import ParametrosSistema


def safe_decimal(val):
    if val is None:
        return Decimal('0')
    try:
        return Decimal(str(val).replace(',', '').replace('$', '').strip())
    except (InvalidOperation, ValueError, TypeError):
        return Decimal('0')


p = ParametrosSistema.objects.filter(activo=True).first()
vigencia = p.vigencia if p else 2026
print(f"Vigencia activa: {vigencia}")

# =============================================
# 1. IMPORTAR ANEXO 2
# =============================================
print("\n=== IMPORTANDO ANEXO 2 ===")
wb = openpyxl.load_workbook('/app/gastos_excel/ANEXO 2 PRESUPUESTO DE GASTOS CENTRAL 2026 (1) (2).xlsx', data_only=True)

sheet_name = None
for name in wb.sheetnames:
    if 'ANEXO' in name.upper() or 'DETALLE' in name.upper():
        sheet_name = name
        break
if not sheet_name:
    sheet_name = wb.sheetnames[0]
print(f"Hoja: {sheet_name}")
ws = wb[sheet_name]

# Detectar columnas
header_row = None
col_map = {}
for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=False), 1):
    for cell in row:
        val = str(cell.value or '').upper().strip()
        if 'SECCION' in val and 'NOMBRE' not in val and 'INVERSION' not in val:
            col_map['seccion'] = cell.column - 1
            header_row = row_idx
        elif 'NOMBRE' in val and 'SECCION' in val:
            col_map['nombre_seccion'] = cell.column - 1
        elif 'IDENTIFICACION' in val or val.startswith('IDENTIFICACI'):
            col_map['codigo'] = cell.column - 1
        elif 'COD' in val and 'FTE' in val:
            if 'cod_fuente' not in col_map:
                col_map['cod_fuente'] = cell.column - 1
        elif 'NOMBRE' in val and 'FUENTE' in val:
            col_map['nombre_fuente'] = cell.column - 1
        elif 'DESCRIPCION' in val or val.startswith('DESCRIPCI'):
            col_map['descripcion'] = cell.column - 1
        elif 'APROPIACION' in val or val.startswith('APROPIACI'):
            col_map['apropiacion'] = cell.column - 1
    if header_row:
        break

print(f"Header row: {header_row}")
print(f"Columnas detectadas: {col_map}")

if not header_row or 'codigo' not in col_map:
    print("ERROR: No se detectaron columnas. Intentando buscar mas...")
    # Intentar buscar en mas filas
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=False), 1):
        for cell in row:
            val = str(cell.value or '').strip()
            if val:
                print(f"  Fila {row_idx}, Col {cell.column}: [{val}]")
    sys.exit(1)

# Limpiar datos previos
RubroGasto.objects.filter(vigencia=vigencia).delete()
SeccionGasto.objects.filter(vigencia=vigencia).delete()
FuenteFinanciacion.objects.filter(vigencia=vigencia).delete()
print("Datos previos eliminados")

count = 0
orden = 0
seccion_cache = {}
fuente_cache = {}
parent_stack = {}

for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
    codigo_raw = ''
    if 'codigo' in col_map and len(row) > col_map['codigo'] and row[col_map['codigo']]:
        codigo_raw = str(row[col_map['codigo']]).strip()

    descripcion = ''
    if 'descripcion' in col_map and len(row) > col_map['descripcion'] and row[col_map['descripcion']]:
        descripcion = str(row[col_map['descripcion']]).strip()

    if not codigo_raw:
        continue

    # Seccion
    seccion_cod = ''
    seccion_nombre = ''
    if 'seccion' in col_map and len(row) > col_map['seccion'] and row[col_map['seccion']]:
        seccion_cod = str(row[col_map['seccion']]).strip()
    if 'nombre_seccion' in col_map and len(row) > col_map['nombre_seccion'] and row[col_map['nombre_seccion']]:
        seccion_nombre = str(row[col_map['nombre_seccion']]).strip()

    seccion_obj = None
    if seccion_cod:
        if seccion_cod not in seccion_cache:
            seccion_obj, _ = SeccionGasto.objects.get_or_create(
                vigencia=vigencia, codigo=seccion_cod,
                defaults={'nombre': seccion_nombre or seccion_cod}
            )
            seccion_cache[seccion_cod] = seccion_obj
        else:
            seccion_obj = seccion_cache[seccion_cod]

    # Fuente
    cod_fuente = ''
    nombre_fuente = ''
    if 'cod_fuente' in col_map and len(row) > col_map['cod_fuente'] and row[col_map['cod_fuente']]:
        cod_fuente = str(row[col_map['cod_fuente']]).strip()
    if 'nombre_fuente' in col_map and len(row) > col_map['nombre_fuente'] and row[col_map['nombre_fuente']]:
        nombre_fuente = str(row[col_map['nombre_fuente']]).strip()

    fuente_obj = None
    if cod_fuente:
        if cod_fuente not in fuente_cache:
            fuente_obj, _ = FuenteFinanciacion.objects.get_or_create(
                vigencia=vigencia, codigo=cod_fuente,
                defaults={'nombre': nombre_fuente or cod_fuente}
            )
            fuente_cache[cod_fuente] = fuente_obj
        else:
            fuente_obj = fuente_cache[cod_fuente]

    # Apropiacion
    val_aprop = Decimal('0')
    if 'apropiacion' in col_map and len(row) > col_map['apropiacion']:
        val_aprop = safe_decimal(row[col_map['apropiacion']])

    # Nivel jerarquico
    parts = codigo_raw.replace('-', '.').split('.')
    nivel = len(parts) - 1
    es_titulo = nivel <= 2

    # Tipo de gasto
    tipo_gasto = 'FUN'
    desc_upper = (descripcion or '').upper()
    cod_upper = codigo_raw.upper()
    if 'INVERSION' in desc_upper or 'INVERSI' in desc_upper:
        tipo_gasto = 'INV'
    elif 'DEUDA' in desc_upper:
        tipo_gasto = 'DEU'

    # Parent
    parent = None
    if nivel > 0:
        parent = parent_stack.get(nivel - 1)

    orden += 1
    rubro = RubroGasto.objects.create(
        vigencia=vigencia,
        codigo=codigo_raw,
        descripcion=descripcion,
        seccion=seccion_obj,
        fuente=fuente_obj,
        codigo_fuente=cod_fuente,
        nombre_fuente=nombre_fuente,
        tipo_gasto=tipo_gasto,
        parent=parent,
        valor_apropiacion=val_aprop if not es_titulo else 0,
        es_titulo=es_titulo,
        orden=orden,
        nivel=nivel,
    )

    if es_titulo:
        parent_stack[nivel] = rubro

    count += 1

# Recalcular titulos
titulos = RubroGasto.objects.filter(vigencia=vigencia, es_titulo=True).order_by('-nivel')
for titulo in titulos:
    titulo.calcular_hijos()

print(f"Anexo 2: {count} rubros importados")
print(f"Secciones creadas: {len(seccion_cache)}")
print(f"Fuentes creadas: {len(fuente_cache)}")

total = RubroGasto.objects.filter(vigencia=vigencia, es_titulo=False).aggregate(
    t=Sum('valor_apropiacion')
)['t'] or 0
print(f"Total apropiacion gastos: ${total:,.0f}")

# =============================================
# 2. IMPORTAR EJECUCION
# =============================================
print("\n=== IMPORTANDO EJECUCION DE GASTOS ===")
wb2 = openpyxl.load_workbook('/app/gastos_excel/EJECUCION GASTOS  FEBRERO 2026 (1).xlsx', data_only=True)

sheet_name2 = None
for name in wb2.sheetnames:
    if 'GASTO' in name.upper():
        sheet_name2 = name
        break
if not sheet_name2:
    sheet_name2 = wb2.sheetnames[0]
print(f"Hoja: {sheet_name2}")
ws2 = wb2[sheet_name2]

# Detectar columnas
header_row2 = None
col_map2 = {}
for row_idx, row in enumerate(ws2.iter_rows(min_row=1, max_row=10, values_only=False), 1):
    for cell in row:
        val = str(cell.value or '').upper().strip()
        if 'RUBRO' in val and 'PPTAL' in val and 'DESCRIPCI' not in val:
            col_map2['codigo'] = cell.column - 1
            header_row2 = row_idx
        elif 'DESCRIPCI' in val and 'RUBRO' in val:
            col_map2['descripcion'] = cell.column - 1
        elif 'PRESUPUESTO' in val and 'APROBADO' in val:
            col_map2['aprobado'] = cell.column - 1
        elif val == 'ADICIONES':
            col_map2['adiciones'] = cell.column - 1
        elif val == 'REDUCCIONES':
            col_map2['reducciones'] = cell.column - 1
        elif 'TRASLADO' in val and ('CREDITO' in val or 'CRÉDITO' in val) and 'CONTRA' not in val:
            col_map2['traslado_credito'] = cell.column - 1
        elif 'CONTRA' in val and ('CREDITO' in val or 'CRÉDITO' in val):
            col_map2['traslado_contra'] = cell.column - 1
        elif 'APLAZAMIENTO' in val and 'DES' not in val:
            col_map2['aplazamientos'] = cell.column - 1
        elif 'DESAPLAZAMIENTO' in val:
            col_map2['desaplazamientos'] = cell.column - 1
        elif 'COMPROMISO' in val:
            col_map2['compromisos'] = cell.column - 1
        elif 'ORDENADO' in val:
            col_map2['ordenado'] = cell.column - 1
        elif 'PAGADO' in val:
            col_map2['pagado'] = cell.column - 1
    if header_row2:
        break

print(f"Header row: {header_row2}")
print(f"Columnas detectadas: {col_map2}")

if header_row2 and 'codigo' in col_map2:
    count_new = 0
    count_updated = 0

    for row in ws2.iter_rows(min_row=header_row2 + 1, values_only=True):
        codigo = ''
        if len(row) > col_map2['codigo'] and row[col_map2['codigo']]:
            codigo = str(row[col_map2['codigo']]).strip()
        if not codigo:
            continue

        descripcion = ''
        if 'descripcion' in col_map2 and len(row) > col_map2['descripcion'] and row[col_map2['descripcion']]:
            descripcion = str(row[col_map2['descripcion']]).strip()

        rubro = RubroGasto.objects.filter(vigencia=vigencia, codigo=codigo).first()
        if not rubro:
            parts = codigo.replace('-', '.').split('.')
            nivel = len(parts) - 1
            rubro = RubroGasto.objects.create(
                vigencia=vigencia, codigo=codigo, descripcion=descripcion,
                nivel=nivel, es_titulo=False,
                orden=RubroGasto.objects.filter(vigencia=vigencia).count() + 1,
            )
            count_new += 1

        ej, created = EjecucionGasto.objects.get_or_create(rubro=rubro)
        if 'aprobado' in col_map2 and len(row) > col_map2['aprobado']:
            ej.presupuesto_aprobado = safe_decimal(row[col_map2['aprobado']])
        if 'adiciones' in col_map2 and len(row) > col_map2['adiciones']:
            ej.adiciones = safe_decimal(row[col_map2['adiciones']])
        if 'reducciones' in col_map2 and len(row) > col_map2['reducciones']:
            ej.reducciones = safe_decimal(row[col_map2['reducciones']])
        if 'traslado_credito' in col_map2 and len(row) > col_map2['traslado_credito']:
            ej.traslado_credito = safe_decimal(row[col_map2['traslado_credito']])
        if 'traslado_contra' in col_map2 and len(row) > col_map2['traslado_contra']:
            ej.traslado_contra_credito = safe_decimal(row[col_map2['traslado_contra']])
        if 'aplazamientos' in col_map2 and len(row) > col_map2['aplazamientos']:
            ej.aplazamientos = safe_decimal(row[col_map2['aplazamientos']])
        if 'desaplazamientos' in col_map2 and len(row) > col_map2['desaplazamientos']:
            ej.desaplazamientos = safe_decimal(row[col_map2['desaplazamientos']])
        if 'compromisos' in col_map2 and len(row) > col_map2['compromisos']:
            ej.compromisos = safe_decimal(row[col_map2['compromisos']])
        if 'ordenado' in col_map2 and len(row) > col_map2['ordenado']:
            ej.ordenado = safe_decimal(row[col_map2['ordenado']])
        if 'pagado' in col_map2 and len(row) > col_map2['pagado']:
            ej.pagado = safe_decimal(row[col_map2['pagado']])
        ej.save()
        count_updated += 1

    print(f"Ejecucion: {count_updated} registros actualizados, {count_new} rubros nuevos")

    totales = EjecucionGasto.objects.filter(rubro__vigencia=vigencia).aggregate(
        total_aprobado=Sum('presupuesto_aprobado'),
        total_compromisos=Sum('compromisos'),
        total_pagado=Sum('pagado'),
    )
    print(f"Total aprobado: ${totales['total_aprobado'] or 0:,.0f}")
    print(f"Total compromisos: ${totales['total_compromisos'] or 0:,.0f}")
    print(f"Total pagado: ${totales['total_pagado'] or 0:,.0f}")
else:
    print("No se pudieron detectar las columnas de ejecucion")
    if not header_row2:
        print("Mostrando encabezados encontrados:")
        for row_idx, row in enumerate(ws2.iter_rows(min_row=1, max_row=5, values_only=False), 1):
            for cell in row:
                val = str(cell.value or '').strip()
                if val:
                    print(f"  Fila {row_idx}, Col {cell.column}: [{val}]")

print("\n=== IMPORTACION COMPLETADA ===")
