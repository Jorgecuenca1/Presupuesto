from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Q, Count
from django.core.paginator import Paginator
from django.http import HttpResponse
from decimal import Decimal, InvalidOperation
import openpyxl

from core.models import ParametrosSistema
from .models import (
    TarifaPredial, CulturaPago, ContribuyentePredial, CarteraVigenciaAnterior,
    TarifaICA, ContribuyenteICA, RubroIngreso, ResumenCalculo,
    CategoriaPredial, ActividadICA, CifraHistoricaIngreso,
)
from .forms import (
    TarifaPredialForm, CulturaPagoForm, ContribuyentePredialForm,
    ImportarExcelForm, CarteraForm, TarifaICAForm, ContribuyenteICAForm,
    RubroIngresoForm, CifraHistoricaIngresoForm,
)
from .utils import (
    calcular_predial, calcular_predial_vigencias_anteriores,
    calcular_ica, calcular_todos_ingresos,
)


def _vigencia():
    p = ParametrosSistema.objects.filter(activo=True).first()
    return p.vigencia if p else 2026


# ─── TARIFAS PREDIAL ──────────────────────────────────────────────
@login_required
def tarifas_predial(request):
    vigencia = _vigencia()
    tarifas = TarifaPredial.objects.filter(vigencia=vigencia)
    culturas = CulturaPago.objects.filter(vigencia=vigencia)
    form_tarifa = TarifaPredialForm(initial={'vigencia': vigencia})
    form_cultura = CulturaPagoForm(initial={'vigencia': vigencia})
    return render(request, 'ingresos/tarifas_predial.html', {
        'tarifas': tarifas, 'culturas': culturas,
        'form_tarifa': form_tarifa, 'form_cultura': form_cultura,
        'vigencia': vigencia,
    })


@login_required
def tarifa_predial_guardar(request):
    if request.method == 'POST':
        pk = request.POST.get('pk')
        instance = get_object_or_404(TarifaPredial, pk=pk) if pk else None
        form = TarifaPredialForm(request.POST, instance=instance)
        if form.is_valid():
            form.save()
            messages.success(request, 'Tarifa guardada')
    return redirect('tarifas_predial')


@login_required
def tarifa_predial_eliminar(request, pk):
    get_object_or_404(TarifaPredial, pk=pk).delete()
    messages.success(request, 'Tarifa eliminada')
    return redirect('tarifas_predial')


@login_required
def cultura_pago_guardar(request):
    if request.method == 'POST':
        pk = request.POST.get('pk')
        instance = get_object_or_404(CulturaPago, pk=pk) if pk else None
        form = CulturaPagoForm(request.POST, instance=instance)
        if form.is_valid():
            form.save()
            messages.success(request, 'Cultura de pago guardada')
    return redirect('tarifas_predial')


# ─── CONTRIBUYENTES PREDIAL ───────────────────────────────────────
@login_required
def contribuyentes_predial(request):
    vigencia = _vigencia()
    q = (request.GET.get('q') or '').strip()
    categoria = (request.GET.get('categoria') or '').strip()
    qs = ContribuyentePredial.objects.filter(vigencia=vigencia)
    if q:
        qs = qs.filter(
            Q(propietario__icontains=q) |
            Q(cedula_catastral__icontains=q) |
            Q(direccion__icontains=q) |
            Q(nombre_predio__icontains=q)
        )
    if categoria:
        qs = qs.filter(categoria=categoria)

    total = qs.count()
    total_avaluo = qs.aggregate(t=Sum('avaluo_catastral'))['t'] or 0

    try:
        per_page = max(10, min(500, int(request.GET.get('per_page') or 50)))
    except ValueError:
        per_page = 50
    paginator = Paginator(qs.order_by('-avaluo_catastral'), per_page)
    page_number = request.GET.get('page') or 1
    page_obj = paginator.get_page(page_number)

    resumen_cat = (ContribuyentePredial.objects.filter(vigencia=vigencia)
                   .values('categoria')
                   .annotate(n=Count('id'), total=Sum('avaluo_catastral'))
                   .order_by('-n'))

    return render(request, 'ingresos/contribuyentes_predial.html', {
        'contribuyentes': page_obj.object_list,
        'page_obj': page_obj,
        'paginator': paginator,
        'total': total,
        'total_avaluo': total_avaluo,
        'vigencia': vigencia,
        'q': q,
        'categoria': categoria,
        'categorias': CategoriaPredial.choices,
        'resumen_cat': resumen_cat,
        'per_page': per_page,
    })


@login_required
def contribuyente_predial_crear(request):
    vigencia = _vigencia()
    if request.method == 'POST':
        form = ContribuyentePredialForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Contribuyente agregado')
            return redirect('contribuyentes_predial')
    else:
        form = ContribuyentePredialForm(initial={'vigencia': vigencia})
    return render(request, 'ingresos/contribuyente_predial_form.html', {'form': form})


@login_required
def contribuyente_predial_eliminar(request, pk):
    get_object_or_404(ContribuyentePredial, pk=pk).delete()
    messages.success(request, 'Contribuyente eliminado')
    return redirect('contribuyentes_predial')


URBANO_TIPOS_SET = {
    # Según clasificación catastral municipal, SOLO la cabecera es urbana.
    # Los corregimientos (Pachaquiaro, Puerto Guadalupe, etc.) son rurales.
    'CABECERA MUNICIPAL',
}


def _clasificar_predio(tipo, destino, clase):
    tipo = (tipo or '').strip().upper()
    destino = (destino or '').strip().upper()
    clase = (clase or '').strip().upper()
    is_urbano = tipo in URBANO_TIPOS_SET
    is_parcelacion = 'PARCELACION' in clase or 'FINCA' in clase
    is_financiero = 'FINANCIERA' in clase or 'FINANCIER' in clase
    is_no_edif_clase = 'NO EDIFICAD' in clase
    is_lote = 'LOTE' in destino
    is_no_urbanizable = 'NO URBANIZABLE' in destino
    is_urbanizable_no_urbanizado = 'URBANIZABLE NO URBAN' in destino
    is_urbanizado_no_edif = destino == 'LOTE URBANIZADO NO CONST'
    is_habitacional = destino == 'HABITACIONAL'
    if is_parcelacion:
        return 'PNE' if is_no_edif_clase else 'PE'
    if is_urbano:
        if is_financiero:
            return 'UEF'
        if is_habitacional:
            return 'UV'
        if is_lote:
            if is_no_urbanizable:
                return 'UNNU'
            if is_urbanizable_no_urbanizado:
                return 'UNEU'
            if is_urbanizado_no_edif:
                return 'UNUE'
            return 'UNEU'
        return 'UED'
    return 'RU'


def _to_decimal(val):
    if val is None or val == '':
        return None
    try:
        if isinstance(val, (int, float, Decimal)):
            return Decimal(str(val))
        return Decimal(str(val).replace(',', '').replace('$', '').strip())
    except (InvalidOperation, ValueError):
        return None


def _detectar_formato_predial(ws):
    """Retorna 'comparativo' si el Excel parece ser la Tabla Predial Comparativo.
    Detecta por: ≥25 columnas y encabezado 'REFERENCIA CATASTRAL' en fila 7 col 1."""
    if ws.max_column < 25:
        return 'simple'
    try:
        val = str(ws.cell(row=7, column=1).value or '').strip().upper()
        if 'REFERENCIA' in val and 'CATASTRAL' in val:
            return 'comparativo'
    except Exception:
        pass
    return 'simple'


def _importar_comparativo(ws, vigencia):
    """Importa formato Tabla Predial Comparativo. Encabezados fila 7, datos desde 8.
    Columnas: 1=ref, 11=propietario, 12=direccion, 13=tipo, 14=destino, 15=clase,
    20=avaluo anterior, 23=avaluo actual."""
    existentes = set(ContribuyentePredial.objects.filter(vigencia=vigencia)
                     .values_list('cedula_catastral', flat=True))
    por_cat = {}
    nuevos = []
    omitidos = 0
    for row in ws.iter_rows(min_row=8, values_only=True):
        if not row or not row[0]:
            continue
        ref = str(row[0]).strip()[:50]
        if ref in existentes:
            omitidos += 1
            continue
        if len(row) < 23:
            continue
        avaluo = _to_decimal(row[22])
        if not avaluo or avaluo <= 0:
            avaluo = _to_decimal(row[19])  # fallback vigencia anterior
        if not avaluo or avaluo <= 0:
            continue
        propietario = str(row[10] or row[2] or '').strip()[:300] or 'SIN PROPIETARIO'
        direccion = str(row[11] or '').strip()[:300]
        tipo = str(row[12] or '').strip()
        destino = str(row[13] or '').strip()
        clase = str(row[14] or '').strip()
        categoria = _clasificar_predio(tipo, destino, clase)
        por_cat[categoria] = por_cat.get(categoria, 0) + 1
        nuevos.append(ContribuyentePredial(
            vigencia=vigencia,
            direccion=direccion or tipo[:300],
            nombre_predio=(destino or clase or 'PREDIO')[:200],
            propietario=propietario,
            cedula_catastral=ref,
            avaluo_catastral=avaluo,
            categoria=categoria,
        ))
        existentes.add(ref)
        if len(nuevos) >= 500:
            ContribuyentePredial.objects.bulk_create(nuevos)
            nuevos = []
    if nuevos:
        ContribuyentePredial.objects.bulk_create(nuevos)
    return por_cat, omitidos


def _importar_simple(ws, vigencia):
    cat_map = {v.lower(): k for k, v in CategoriaPredial.choices}
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 5:
            continue
        direccion = str(row[0] or '').strip()
        nombre = str(row[1] or '').strip()
        propietario = str(row[2] or '').strip()
        avaluo = _to_decimal(row[3])
        if not avaluo:
            continue
        cat_raw = str(row[4] or '').strip()
        cedula_cat = str(row[5]).strip() if len(row) > 5 and row[5] else ''
        categoria = cat_map.get(cat_raw.lower(), cat_raw.upper()[:4])
        if categoria not in dict(CategoriaPredial.choices):
            categoria = 'UV'
        ContribuyentePredial.objects.create(
            vigencia=vigencia, direccion=direccion, nombre_predio=nombre,
            propietario=propietario, avaluo_catastral=avaluo,
            categoria=categoria, cedula_catastral=cedula_cat,
        )
        count += 1
    return count


@login_required
def importar_predial(request):
    if request.method == 'POST':
        form = ImportarExcelForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES['archivo']
            vigencia = _vigencia()
            try:
                wb = openpyxl.load_workbook(archivo, data_only=True, read_only=True)
                ws = wb.active
                formato = _detectar_formato_predial(ws)
                if formato == 'comparativo':
                    por_cat, omitidos = _importar_comparativo(ws, vigencia)
                    total = sum(por_cat.values())
                    detalle = ', '.join(f'{k}={v}' for k, v in sorted(por_cat.items(), key=lambda x: -x[1]))
                    messages.success(
                        request,
                        f'{total} predios importados (Tabla Comparativo). {omitidos} ya existían. Por categoría: {detalle}'
                    )
                else:
                    count = _importar_simple(ws, vigencia)
                    messages.success(request, f'{count} contribuyentes importados')
                return redirect('contribuyentes_predial')
            except Exception as e:
                messages.error(request, f'Error al importar: {e}')
    else:
        form = ImportarExcelForm()
    return render(request, 'ingresos/importar_contribuyentes.html', {
        'form': form, 'tipo': 'Predial',
        'columnas': [
            'Dirección', 'Nombre Predio', 'Propietario',
            'Avalúo Catastral', 'Categoría', 'Cédula Catastral (opcional)',
        ],
        'categorias': CategoriaPredial.choices,
        'info_extra': (
            'También acepta el formato "Tabla Predial Comparativo con vigencia anterior" '
            '(32 columnas, encabezados en fila 7). En ese caso el sistema clasifica '
            'automáticamente cada predio por TIPO/DESTINO/CLASE y usa el Avalúo 2026 '
            '(columna 23) con fallback al avalúo anterior (columna 20).'
        ),
    })


# ─── CÁLCULO PREDIAL ──────────────────────────────────────────────
@login_required
def calculo_predial(request):
    vigencia = _vigencia()
    params = ParametrosSistema.objects.filter(vigencia=vigencia).first()
    if request.method == 'POST':
        calcular_predial(vigencia, 'urbano')
        calcular_predial(vigencia, 'rural')
        messages.success(request, 'Cálculo predial ejecutado correctamente')
        return redirect('calculo_predial')

    resumen_urbano = ResumenCalculo.objects.filter(vigencia=vigencia, tipo='predial_urbano')
    resumen_rural = ResumenCalculo.objects.filter(vigencia=vigencia, tipo='predial_rural')
    agg_urb = resumen_urbano.aggregate(
        proy=Sum('proyeccion'), avaluo=Sum('total_avaluo'),
        recaudo=Sum('recaudo_potencial'), predios=Sum('cantidad_predios'),
    )
    agg_rur = resumen_rural.aggregate(
        proy=Sum('proyeccion'), avaluo=Sum('total_avaluo'),
        recaudo=Sum('recaudo_potencial'), predios=Sum('cantidad_predios'),
    )
    total_urbano = agg_urb['proy'] or 0
    total_rural = agg_rur['proy'] or 0
    total_avaluo_urb = agg_urb['avaluo'] or 0
    total_avaluo_rur = agg_rur['avaluo'] or 0
    total_recaudo_urb = agg_urb['recaudo'] or 0
    total_recaudo_rur = agg_rur['recaudo'] or 0
    total_predios_urb = agg_urb['predios'] or 0
    total_predios_rur = agg_rur['predios'] or 0

    total_urb_ant, detalle_urb_ant = calcular_predial_vigencias_anteriores(vigencia, 'urbano')
    total_rur_ant, detalle_rur_ant = calcular_predial_vigencias_anteriores(vigencia, 'rural')

    carteras = CarteraVigenciaAnterior.objects.filter(vigencia_calculo=vigencia)
    form_cartera = CarteraForm(initial={'vigencia_calculo': vigencia})

    return render(request, 'ingresos/calculo_predial.html', {
        'params': params,
        'resumen_urbano': resumen_urbano, 'total_urbano': total_urbano,
        'resumen_rural': resumen_rural, 'total_rural': total_rural,
        'total_avaluo_urb': total_avaluo_urb, 'total_avaluo_rur': total_avaluo_rur,
        'total_recaudo_urb': total_recaudo_urb, 'total_recaudo_rur': total_recaudo_rur,
        'total_predios_urb': total_predios_urb, 'total_predios_rur': total_predios_rur,
        'total_urb_ant': total_urb_ant, 'detalle_urb_ant': detalle_urb_ant,
        'total_rur_ant': total_rur_ant, 'detalle_rur_ant': detalle_rur_ant,
        'carteras': carteras, 'form_cartera': form_cartera,
        'total_predial': total_urbano + total_rural + total_urb_ant + total_rur_ant,
    })


@login_required
def cartera_guardar(request):
    if request.method == 'POST':
        pk = request.POST.get('pk')
        instance = get_object_or_404(CarteraVigenciaAnterior, pk=pk) if pk else None
        form = CarteraForm(request.POST, instance=instance)
        if form.is_valid():
            form.save()
            messages.success(request, 'Cartera guardada')
    return redirect('calculo_predial')


@login_required
def cartera_eliminar(request, pk):
    get_object_or_404(CarteraVigenciaAnterior, pk=pk).delete()
    messages.success(request, 'Cartera eliminada')
    return redirect('calculo_predial')


# ─── TARIFAS ICA ──────────────────────────────────────────────────
@login_required
def tarifas_ica(request):
    vigencia = _vigencia()
    tarifas = TarifaICA.objects.filter(vigencia=vigencia)
    form = TarifaICAForm(initial={'vigencia': vigencia})
    return render(request, 'ingresos/tarifas_ica.html', {
        'tarifas': tarifas, 'form': form, 'vigencia': vigencia,
    })


@login_required
def tarifa_ica_guardar(request):
    if request.method == 'POST':
        pk = request.POST.get('pk')
        instance = get_object_or_404(TarifaICA, pk=pk) if pk else None
        form = TarifaICAForm(request.POST, instance=instance)
        if form.is_valid():
            form.save()
            messages.success(request, 'Tarifa ICA guardada')
    return redirect('tarifas_ica')


@login_required
def tarifa_ica_eliminar(request, pk):
    get_object_or_404(TarifaICA, pk=pk).delete()
    messages.success(request, 'Tarifa eliminada')
    return redirect('tarifas_ica')


# ─── CONTRIBUYENTES ICA ───────────────────────────────────────────
@login_required
def contribuyentes_ica(request):
    vigencia = _vigencia()
    contribuyentes = ContribuyenteICA.objects.filter(vigencia=vigencia)
    total = contribuyentes.count()
    total_ingresos = contribuyentes.aggregate(t=Sum('ingresos_brutos'))['t'] or 0
    return render(request, 'ingresos/contribuyentes_ica.html', {
        'contribuyentes': contribuyentes[:200],
        'total': total, 'total_ingresos': total_ingresos,
        'vigencia': vigencia,
    })


@login_required
def contribuyente_ica_crear(request):
    vigencia = _vigencia()
    if request.method == 'POST':
        form = ContribuyenteICAForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Contribuyente ICA agregado')
            return redirect('contribuyentes_ica')
    else:
        form = ContribuyenteICAForm(initial={'vigencia': vigencia})
    return render(request, 'ingresos/contribuyente_ica_form.html', {'form': form})


@login_required
def contribuyente_ica_eliminar(request, pk):
    get_object_or_404(ContribuyenteICA, pk=pk).delete()
    messages.success(request, 'Contribuyente eliminado')
    return redirect('contribuyentes_ica')


@login_required
def importar_ica(request):
    if request.method == 'POST':
        form = ImportarExcelForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES['archivo']
            vigencia = _vigencia()
            try:
                wb = openpyxl.load_workbook(archivo)
                ws = wb.active
                count = 0
                act_map = {v.lower(): k for k, v in ActividadICA.choices}
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or len(row) < 4:
                        continue
                    nombre = str(row[0] or '').strip()
                    nit = str(row[1] or '').strip()
                    act_raw = str(row[2] or '').strip()
                    try:
                        ingresos = Decimal(str(row[3]).replace(',', '').replace('$', '').strip())
                    except (ValueError, TypeError):
                        continue
                    actividad = act_map.get(act_raw.lower(), act_raw[:3])
                    if actividad not in dict(ActividadICA.choices):
                        actividad = '204'
                    ContribuyenteICA.objects.create(
                        vigencia=vigencia, nombre=nombre, nit=nit,
                        actividad=actividad, ingresos_brutos=ingresos,
                    )
                    count += 1
                messages.success(request, f'{count} contribuyentes ICA importados')
                return redirect('contribuyentes_ica')
            except Exception as e:
                messages.error(request, f'Error al importar: {e}')
    else:
        form = ImportarExcelForm()
    return render(request, 'ingresos/importar_contribuyentes.html', {
        'form': form, 'tipo': 'Industria y Comercio',
        'columnas': ['Nombre/Razón Social', 'NIT', 'Código Actividad (101/201/202/203/204/301/302/401)', 'Ingresos Brutos ($)'],
        'categorias': ActividadICA.choices,
    })


# ─── CÁLCULO ICA ──────────────────────────────────────────────────
@login_required
def calculo_ica(request):
    vigencia = _vigencia()
    params = ParametrosSistema.objects.filter(vigencia=vigencia).first()
    if request.method == 'POST':
        calcular_ica(vigencia)
        messages.success(request, 'Cálculo ICA ejecutado correctamente')
        return redirect('calculo_ica')

    resumen = ResumenCalculo.objects.filter(vigencia=vigencia, tipo='ica')
    total_ica = resumen.aggregate(t=Sum('proyeccion'))['t'] or 0
    avisos_tableros = total_ica * Decimal('0.15')

    contribuyentes = ContribuyenteICA.objects.filter(vigencia=vigencia).order_by('actividad')

    return render(request, 'ingresos/calculo_ica.html', {
        'params': params,
        'resumen': resumen,
        'total_ica': total_ica,
        'avisos_tableros': avisos_tableros,
        'contribuyentes': contribuyentes,
    })


# ─── RUBROS DE INGRESO ────────────────────────────────────────────
@login_required
def rubros_list(request):
    vigencia = _vigencia()
    rubros = RubroIngreso.objects.filter(vigencia=vigencia)
    total = rubros.filter(es_titulo=False).aggregate(t=Sum('valor_apropiacion'))['t'] or 0
    return render(request, 'ingresos/rubros_list.html', {
        'rubros': rubros, 'vigencia': vigencia, 'total': total,
    })


@login_required
def rubro_crear(request):
    vigencia = _vigencia()
    if request.method == 'POST':
        form = RubroIngresoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Rubro creado')
            return redirect('rubros_list')
    else:
        form = RubroIngresoForm(initial={'vigencia': vigencia})
    return render(request, 'ingresos/rubro_form.html', {'form': form, 'titulo': 'Crear Rubro'})


@login_required
def rubro_editar(request, pk):
    rubro = get_object_or_404(RubroIngreso, pk=pk)
    if request.method == 'POST':
        form = RubroIngresoForm(request.POST, instance=rubro)
        if form.is_valid():
            form.save()
            messages.success(request, 'Rubro actualizado')
            return redirect('rubros_list')
    else:
        form = RubroIngresoForm(instance=rubro)
    return render(request, 'ingresos/rubro_form.html', {'form': form, 'titulo': 'Editar Rubro'})


@login_required
def rubro_eliminar(request, pk):
    get_object_or_404(RubroIngreso, pk=pk).delete()
    messages.success(request, 'Rubro eliminado')
    return redirect('rubros_list')


# ─── CALCULAR TODOS ───────────────────────────────────────────────
@login_required
def calcular_todos(request):
    vigencia = _vigencia()
    if request.method == 'POST':
        calcular_todos_ingresos(vigencia)
        messages.success(request, 'Todos los ingresos fueron calculados correctamente')
    return redirect('reporte_ingresos')


# ─── REPORTE ──────────────────────────────────────────────────────
@login_required
def reporte_ingresos(request):
    vigencia = _vigencia()
    params = ParametrosSistema.objects.filter(vigencia=vigencia).first()
    rubros = RubroIngreso.objects.filter(vigencia=vigencia)
    total = rubros.filter(nivel=0, es_titulo=True).aggregate(t=Sum('valor_apropiacion'))['t']
    if not total:
        total = rubros.filter(es_titulo=False).aggregate(t=Sum('valor_apropiacion'))['t'] or 0
    return render(request, 'ingresos/reporte_ingresos.html', {
        'rubros': rubros, 'vigencia': vigencia, 'params': params, 'total': total,
    })


@login_required
def exportar_reporte_excel(request):
    vigencia = _vigencia()
    rubros = RubroIngreso.objects.filter(vigencia=vigencia)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Anexo 1 Detalle Ingresos'
    ws.append(['MUNICIPIO DE PUERTO LÓPEZ'])
    ws.append(['PRESUPUESTO DE INGRESOS Y RECURSOS DE CAPITAL'])
    ws.append([f'VIGENCIA FISCAL {vigencia}'])
    ws.append([])
    ws.append(['CÓDIGO', 'FUENTE', 'DESCRIPCIÓN', 'APROPIACIÓN', 'OBSERVACIONES'])

    for rubro in rubros:
        indent = '  ' * rubro.nivel
        ws.append([
            rubro.codigo,
            rubro.codigo_fuente,
            indent + rubro.descripcion,
            float(rubro.valor_apropiacion),
            rubro.observaciones,
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=Anexo1_Ingresos_{vigencia}.xlsx'
    wb.save(response)
    return response


# ─── CIFRAS HISTÓRICAS INGRESOS ──────────────────────────────────
def _calcular_tcpa(valores_por_anio):
    """Calcula la Tasa Compuesta Promedio Anual.
    TCPA = (Vn/V0)^(1/n) - 1
    valores_por_anio: dict {año: valor_total}
    """
    if len(valores_por_anio) < 2:
        return Decimal('0')
    anios = sorted(valores_por_anio.keys())
    v0 = valores_por_anio[anios[0]]
    vn = valores_por_anio[anios[-1]]
    n = len(anios) - 1
    if v0 <= 0 or vn <= 0:
        return Decimal('0')
    ratio = float(vn) / float(v0)
    tcpa = ratio ** (1.0 / n) - 1.0
    return Decimal(str(round(tcpa, 6)))


@login_required
def cifras_historicas_ingresos(request):
    vigencia = _vigencia()
    cifras = CifraHistoricaIngreso.objects.filter(vigencia_calculo=vigencia)
    anios = sorted(set(cifras.values_list('anio', flat=True)))

    # Totales por año
    totales_por_anio = {}
    icld_por_anio = {}
    icld_sin_sgp_por_anio = {}
    icld_total_por_anio = {}
    for anio in anios:
        cifras_anio = cifras.filter(anio=anio)
        totales_por_anio[anio] = cifras_anio.aggregate(t=Sum('valor_recaudo'))['t'] or Decimal('0')
        icld_por_anio[anio] = cifras_anio.filter(es_icld=True).aggregate(t=Sum('valor_recaudo'))['t'] or Decimal('0')
        sgp_libre = cifras_anio.filter(es_sgp_libre=True).aggregate(t=Sum('valor_recaudo'))['t'] or Decimal('0')
        icld_sin_sgp_por_anio[anio] = icld_por_anio[anio]
        icld_total_por_anio[anio] = icld_por_anio[anio] + sgp_libre

    tcpa = _calcular_tcpa(totales_por_anio)
    tcpa_icld = _calcular_tcpa(icld_por_anio)

    # ICLD netos sin SGP (último año) para cálculo 1% Ley 99/93
    ultimo_anio = anios[-1] if anios else vigencia - 1
    icld_netos_sin_sgp = icld_sin_sgp_por_anio.get(ultimo_anio, Decimal('0'))
    valor_medio_ambiente = icld_netos_sin_sgp * Decimal('0.01')  # 1% Ley 99/93

    # ICLD totales con SGP libre (último año) para indicador Ley 617
    icld_totales = icld_total_por_anio.get(ultimo_anio, Decimal('0'))

    # Rubros únicos para tabla cruzada
    rubros_unicos = list(cifras.values('codigo_rubro', 'descripcion', 'es_icld', 'es_sgp', 'es_sgp_libre')
                         .distinct().order_by('codigo_rubro'))
    for rubro in rubros_unicos:
        rubro['valores'] = {}
        for anio in anios:
            cifra = cifras.filter(anio=anio, codigo_rubro=rubro['codigo_rubro']).first()
            rubro['valores'][anio] = cifra.valor_recaudo if cifra else Decimal('0')

    form = CifraHistoricaIngresoForm(initial={'vigencia_calculo': vigencia})

    return render(request, 'ingresos/cifras_historicas.html', {
        'cifras': cifras,
        'anios': anios,
        'rubros_unicos': rubros_unicos,
        'totales_por_anio': totales_por_anio,
        'icld_por_anio': icld_por_anio,
        'icld_sin_sgp_por_anio': icld_sin_sgp_por_anio,
        'icld_total_por_anio': icld_total_por_anio,
        'tcpa': tcpa,
        'tcpa_icld': tcpa_icld,
        'icld_netos_sin_sgp': icld_netos_sin_sgp,
        'valor_medio_ambiente': valor_medio_ambiente,
        'icld_totales': icld_totales,
        'vigencia': vigencia,
        'form': form,
    })


@login_required
def cifra_historica_ingreso_guardar(request):
    if request.method == 'POST':
        pk = request.POST.get('pk')
        instance = get_object_or_404(CifraHistoricaIngreso, pk=pk) if pk else None
        form = CifraHistoricaIngresoForm(request.POST, instance=instance)
        if form.is_valid():
            form.save()
            messages.success(request, 'Cifra histórica guardada')
    return redirect('cifras_historicas_ingresos')


@login_required
def cifra_historica_ingreso_eliminar(request, pk):
    get_object_or_404(CifraHistoricaIngreso, pk=pk).delete()
    messages.success(request, 'Cifra histórica eliminada')
    return redirect('cifras_historicas_ingresos')


@login_required
def importar_cifras_historicas_ingresos(request):
    """Importa cifras históricas desde Excel. Columnas: Código, Descripción, 2022, 2023, 2024, 2025"""
    if request.method == 'POST':
        form = ImportarExcelForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES['archivo']
            vigencia = _vigencia()
            try:
                wb = openpyxl.load_workbook(archivo, data_only=True)
                ws = wb.active
                # Primera fila = encabezados, detectar años
                header = [str(c.value or '').strip() for c in ws[1]]
                anio_cols = {}
                for idx, h in enumerate(header):
                    try:
                        anio = int(h)
                        if 2020 <= anio <= 2030:
                            anio_cols[anio] = idx
                    except ValueError:
                        pass
                if not anio_cols:
                    messages.error(request, 'No se encontraron columnas de años (2022, 2023, etc.)')
                    return redirect('cifras_historicas_ingresos')

                count = 0
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or not row[0]:
                        continue
                    codigo = str(row[0]).strip()
                    descripcion = str(row[1] or '').strip() if len(row) > 1 else ''
                    es_icld = False
                    es_sgp = False
                    es_sgp_libre = False
                    # Detectar flags si hay columnas adicionales
                    for idx, h in enumerate(header):
                        h_upper = h.upper()
                        if 'ICLD' in h_upper and len(row) > idx:
                            val = str(row[idx] or '').upper().strip()
                            es_icld = val in ('SI', 'SÍ', 'S', '1', 'X', 'TRUE')
                        elif 'SGP LIBRE' in h_upper and len(row) > idx:
                            val = str(row[idx] or '').upper().strip()
                            es_sgp_libre = val in ('SI', 'SÍ', 'S', '1', 'X', 'TRUE')
                        elif 'SGP' in h_upper and 'LIBRE' not in h_upper and len(row) > idx:
                            val = str(row[idx] or '').upper().strip()
                            es_sgp = val in ('SI', 'SÍ', 'S', '1', 'X', 'TRUE')

                    for anio, col_idx in anio_cols.items():
                        if col_idx < len(row) and row[col_idx]:
                            try:
                                valor = Decimal(str(row[col_idx]).replace(',', '').replace('$', '').strip())
                            except Exception:
                                continue
                            CifraHistoricaIngreso.objects.update_or_create(
                                vigencia_calculo=vigencia, anio=anio, codigo_rubro=codigo,
                                defaults={
                                    'descripcion': descripcion,
                                    'valor_recaudo': valor,
                                    'es_icld': es_icld,
                                    'es_sgp': es_sgp,
                                    'es_sgp_libre': es_sgp_libre,
                                }
                            )
                            count += 1

                messages.success(request, f'{count} cifras históricas importadas')
                return redirect('cifras_historicas_ingresos')
            except Exception as e:
                messages.error(request, f'Error al importar: {e}')
    else:
        form = ImportarExcelForm()
    return render(request, 'ingresos/importar_contribuyentes.html', {
        'form': form, 'tipo': 'Cifras Históricas Ingresos CUIPO',
        'columnas': ['Código', 'Descripción', '2022', '2023', '2024', '2025', 'ICLD (SI/NO)', 'SGP (SI/NO)', 'SGP Libre (SI/NO)'],
        'categorias': [],
    })


@login_required
def calcular_tcpa_ingresos(request):
    """Calcula TCPA y actualiza parámetros del sistema"""
    vigencia = _vigencia()
    cifras = CifraHistoricaIngreso.objects.filter(vigencia_calculo=vigencia)
    anios = sorted(set(cifras.values_list('anio', flat=True)))
    totales_por_anio = {}
    for anio in anios:
        totales_por_anio[anio] = cifras.filter(anio=anio).aggregate(t=Sum('valor_recaudo'))['t'] or Decimal('0')

    tcpa = _calcular_tcpa(totales_por_anio)
    params = ParametrosSistema.objects.filter(vigencia=vigencia).first()
    if params:
        params.tcpa_ingresos = tcpa
        params.save(update_fields=['tcpa_ingresos'])
        messages.success(request, f'TCPA Ingresos calculada: {tcpa * 100:.2f}%')
    else:
        messages.error(request, 'No hay parámetros configurados para la vigencia')
    return redirect('cifras_historicas_ingresos')
