"""Importa Sistema_MFMP_PuertoLopez_v75.xlsx a la BD (modelos MFMP nuevos).

Uso:
    python import_mfmp_v75.py [/ruta/al/xlsx]

Carga:
    - Fuentes (catálogo maestro)
    - Plan Financiero (10 años A/B/C/D)
    - ICLD proyectado 10 años por fuente
    - Ley 617 proyectada
    - POAI 10 años por fuente
    - POAI por dependencia
    - Cuadre por fuente
    - Saldo VF por fuente
    - Ingresos corrientes Ley 358
"""
import os, sys, warnings
warnings.filterwarnings('ignore')
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'presupuesto_project.settings')
import django; django.setup()
from decimal import Decimal, InvalidOperation
from openpyxl import load_workbook
from core.models import (
    FuenteFinanciacion, PlanFinancieroLinea, ICLDProyectado,
    Ley617Proyectado, POAIProyectado, POAIPorDependencia,
    CuadrePorFuente, SaldoVFPorFuente, IngresoCorrienteLey358,
    Refinanciacion, RefinanciacionProyeccion,
    CCPETIngreso, CCPETGasto, ParametrosSistema,
    ProyeccionRubroIngreso, ProyeccionRubroGasto,
    CargaPOAIProyecto, ICOProyeccion, PlantaDetalleCargo,
    ParametroAnualPredial, ParametroAnualPlanta, BaseEstampillasAnual,
)


def D(v):
    if v is None or v == '':
        return Decimal('0')
    try:
        return Decimal(str(v))
    except (InvalidOperation, ValueError):
        return Decimal('0')


def importar(xlsx_path):
    wb = load_workbook(xlsx_path, data_only=True)

    # ═══ 1) FUENTES (catálogo) ═══════════════════════════════════════════
    ws = wb['Fuentes']
    n = 0
    for r in range(2, ws.max_row + 1):
        cod = ws.cell(row=r, column=1).value
        nom = ws.cell(row=r, column=2).value
        if not cod or not nom:
            continue
        FuenteFinanciacion.objects.update_or_create(
            codigo=str(cod).strip(),
            defaults={'nombre': str(nom).strip(), 'activo': True},
        )
        n += 1
    print(f'✓ Fuentes:                       {n:>5}')

    # ═══ 2) PLAN FINANCIERO ══════════════════════════════════════════════
    ws = wb['Plan Financiero']
    # Headers en F2: cols 2..12 = 2026..2036
    anios = [ws.cell(row=2, column=c).value for c in range(2, 13)]
    anios = [int(a) for a in anios if a]
    MAPA = {
        'A. INGRESOS TOTALES': 'A',
        'B. FUNCIONAMIENTO': 'B',
        '   B.1 Personal': 'B1',
        '   B.2 Bienes, Servicios y Otros': 'B2',
        'C. SERVICIO DE LA DEUDA': 'C',
        'D. INVERSIÓN (= A − B − C)': 'D',
        'Total Gastos (B+C+D)': 'T',
    }
    n = 0
    for r in range(3, ws.max_row + 1):
        concepto = ws.cell(row=r, column=1).value
        if not concepto:
            continue
        tipo = MAPA.get(str(concepto))
        if not tipo:
            # tolerar espacios/prefijos leves
            for k, v in MAPA.items():
                if str(concepto).strip().startswith(k.strip()[:20]):
                    tipo = v; break
        if not tipo:
            continue
        for i, anio in enumerate(anios):
            val = D(ws.cell(row=r, column=2 + i).value)
            PlanFinancieroLinea.objects.update_or_create(
                tipo=tipo, anio=anio, defaults={'valor': val},
            )
            n += 1
    print(f'✓ Plan Financiero (celdas):      {n:>5}')

    # ═══ 3) ICLD PROYECTADO ══════════════════════════════════════════════
    ws = wb['ICLD']
    # Headers están en F10 (2026-2036 desde col 3)
    row_h = 10
    anios = [ws.cell(row=row_h, column=c).value for c in range(3, 14)]
    anios = [int(a) for a in anios if a]
    n = 0
    for r in range(11, ws.max_row + 1):
        concepto = ws.cell(row=r, column=2).value
        if not concepto:
            continue
        concepto_s = str(concepto).strip().upper()
        # Solo capturamos "Total Fuente N (...)"
        if 'TOTAL FUENTE' not in concepto_s:
            continue
        # Extraer código
        try:
            cod = concepto_s.split('FUENTE')[1].split('(')[0].strip()
        except Exception:
            continue
        fuente = FuenteFinanciacion.objects.filter(codigo=cod).first()
        if not fuente:
            continue
        for i, anio in enumerate(anios):
            val = D(ws.cell(row=r, column=3 + i).value)
            ICLDProyectado.objects.update_or_create(
                fuente=fuente, anio=anio,
                defaults={'valor_bruto': val},
            )
            n += 1
    print(f'✓ ICLD Proyectado (celdas):      {n:>5}')

    # ═══ 4) LEY 617 PROYECTADO ═══════════════════════════════════════════
    ws = wb['Ley 617']
    # Headers en F8 (col 3..13 = 2026..2036)
    row_h = 8
    anios = [ws.cell(row=row_h, column=c).value for c in range(3, 14)]
    anios = [int(a) for a in anios if a]
    # F9: Gastos de Funcionamiento, F10: ICLD Neto (buscar solo primer match)
    gf_row = None; icld_row = None; pct_row = None
    for r in range(9, min(ws.max_row + 1, 14)):
        c = ws.cell(row=r, column=2).value
        if not c: continue
        cs = str(c).lower()
        # Excluir filas de indicador/límite que también contienen palabras clave
        if 'indicador' in cs or 'cumple' in cs:
            if 'limite' in cs or 'límite' in cs:
                pct_row = r
            continue
        if gf_row is None and 'gasto' in cs and 'funcion' in cs: gf_row = r
        if icld_row is None and 'icld' in cs and 'neto' in cs: icld_row = r
    if gf_row and icld_row:
        n = 0
        for i, anio in enumerate(anios):
            gf = D(ws.cell(row=gf_row, column=3 + i).value)
            icld = D(ws.cell(row=icld_row, column=3 + i).value)
            pct = D(ws.cell(row=pct_row, column=3 + i).value) if pct_row else Decimal('80')
            Ley617Proyectado.objects.update_or_create(
                anio=anio,
                defaults={'gastos_funcionamiento': gf, 'icld_neto': icld,
                          'pct_limite': pct if pct > 0 else Decimal('80')},
            )
            n += 1
        print(f'✓ Ley 617 Proyectado (años):     {n:>5}')

    # ═══ 5) POAI 2027-2036 por fuente ════════════════════════════════════
    ws = wb['POAI 2027-2036']
    # Headers en F2 col 3..12 = 2027..2036
    anios = [ws.cell(row=2, column=c).value for c in range(3, 13)]
    anios = [int(a) for a in anios if a]
    n = 0
    for r in range(3, ws.max_row + 1):
        cod = ws.cell(row=r, column=1).value
        if not cod: continue
        fuente = FuenteFinanciacion.objects.filter(codigo=str(cod).strip()).first()
        if not fuente: continue
        for i, anio in enumerate(anios):
            val = D(ws.cell(row=r, column=3 + i).value)
            if val == 0: continue
            POAIProyectado.objects.update_or_create(
                fuente=fuente, anio=anio,
                defaults={'valor': val},
            )
            n += 1
    print(f'✓ POAI Proyectado (celdas):      {n:>5}')

    # ═══ 6) POAI x DEPENDENCIA ═══════════════════════════════════════════
    ws = wb['POAI x Dependencia']
    anios = [ws.cell(row=2, column=c).value for c in range(3, 13)]
    anios = [int(a) for a in anios if a]
    n = 0
    for r in range(3, ws.max_row + 1):
        dep = ws.cell(row=r, column=1).value
        pct = ws.cell(row=r, column=2).value
        if not dep: continue
        for i, anio in enumerate(anios):
            val = D(ws.cell(row=r, column=3 + i).value)
            POAIPorDependencia.objects.update_or_create(
                dependencia=str(dep).strip(), anio=anio,
                defaults={'pct_participacion': D(pct), 'valor': val},
            )
            n += 1
    print(f'✓ POAI Dependencias (celdas):    {n:>5}')

    # ═══ 7) CUADRE POR FUENTE ════════════════════════════════════════════
    ws = wb['Cuadre por Fuente']
    # F3 headers: cols 3=Ing 2026, 4=Gto 2026, 6=Ing 2027, 7=Gto 2027
    n = 0
    for r in range(4, ws.max_row + 1):
        cod = ws.cell(row=r, column=1).value
        if not cod: continue
        fuente = FuenteFinanciacion.objects.filter(codigo=str(cod).strip()).first()
        if not fuente: continue
        # 2026
        ing_26 = D(ws.cell(row=r, column=3).value)
        gto_26 = D(ws.cell(row=r, column=4).value)
        CuadrePorFuente.objects.update_or_create(
            fuente=fuente, anio=2026,
            defaults={'ingreso': ing_26, 'gasto': gto_26},
        )
        # 2027
        ing_27 = D(ws.cell(row=r, column=6).value)
        gto_27 = D(ws.cell(row=r, column=7).value)
        CuadrePorFuente.objects.update_or_create(
            fuente=fuente, anio=2027,
            defaults={'ingreso': ing_27, 'gasto': gto_27},
        )
        n += 2
    print(f'✓ Cuadre por Fuente (registros): {n:>5}')

    # ═══ 8) SALDO VF POR FUENTE ══════════════════════════════════════════
    ws = wb['Saldo VF por Fuente']
    # F3 headers
    anio_ref = ws.cell(row=1, column=14).value
    try: anio_ref = int(anio_ref)
    except (TypeError, ValueError): anio_ref = 2026
    n = 0
    for r in range(4, ws.max_row + 1):
        cod = ws.cell(row=r, column=1).value
        if not cod: continue
        fuente = FuenteFinanciacion.objects.filter(codigo=str(cod).strip()).first()
        if not fuente: continue
        SaldoVFPorFuente.objects.update_or_create(
            fuente=fuente, anio_referencia=anio_ref,
            defaults={
                'apropiacion_definitiva': D(ws.cell(row=r, column=4).value),
                'cdp_vigentes': D(ws.cell(row=r, column=6).value),
                'vf_aprobadas': D(ws.cell(row=r, column=7).value),
                'vf_en_tramite': D(ws.cell(row=r, column=8).value),
                'vf_solicitada': D(ws.cell(row=r, column=10).value),
            },
        )
        n += 1
    print(f'✓ Saldo VF por Fuente:           {n:>5}')

    # ═══ 9) INGRESOS CORRIENTES LEY 358 ══════════════════════════════════
    ws = wb['Ing Corrientes Ley 358']
    n = 0
    for r in range(4, ws.max_row + 1):
        cod = ws.cell(row=r, column=1).value
        desc = ws.cell(row=r, column=2).value
        fte = ws.cell(row=r, column=3).value
        if not cod or not desc: continue
        IngresoCorrienteLey358.objects.update_or_create(
            codigo_ccpet=str(cod).strip(),
            fuente=str(fte or '').strip(),
            defaults={
                'descripcion': str(desc).strip()[:300],
                'ejec_2024': D(ws.cell(row=r, column=4).value),
                'ejec_2025': D(ws.cell(row=r, column=5).value),
                'aforo_2026': D(ws.cell(row=r, column=6).value),
                'aplica_ley_358': bool(int(D(ws.cell(row=r, column=7).value))),
            },
        )
        n += 1
    print(f'✓ Ing Corrientes Ley 358:        {n:>5}')

    # ═══ 10) REFINANCIACIÓN ══════════════════════════════════════════════
    ws = wb['Refinanciacion']
    # Header parámetros en F3-F9
    aplicar = bool(int(D(ws.cell(row=3, column=2).value)))
    anio_ref = int(D(ws.cell(row=4, column=2).value))
    tasa = D(ws.cell(row=5, column=2).value)
    plazo = int(D(ws.cell(row=6, column=2).value))
    gracia = int(D(ws.cell(row=7, column=2).value))
    saldo = D(ws.cell(row=8, column=2).value) * Decimal('1000000')  # viene en millones
    pagare_obj = str(ws.cell(row=9, column=2).value or 'Pagaré 1 - BBVA')
    r_obj, _ = Refinanciacion.objects.update_or_create(
        pk=1, defaults={
            'aplicar': aplicar, 'anio_refinanciacion': anio_ref,
            'nueva_tasa_ea': tasa, 'nuevo_plazo_anios': plazo,
            'anios_gracia': gracia, 'saldo_refinanciar': saldo,
            'pagare_objetivo': pagare_obj[:100],
        }
    )
    # Años en F10: cols 3..14 = 2025..2036
    anios = [ws.cell(row=10, column=c).value for c in range(3, 15)]
    anios = [int(a) for a in anios if a]
    n = 0
    for i, anio in enumerate(anios):
        col = 3 + i
        saldo_o = D(ws.cell(row=11, column=col).value) * Decimal('1000000')
        int_o = D(ws.cell(row=12, column=col).value) * Decimal('1000000')
        amort_o = D(ws.cell(row=13, column=col).value) * Decimal('1000000')
        RefinanciacionProyeccion.objects.update_or_create(
            anio=anio, defaults={
                'saldo_original': saldo_o,
                'intereses': int_o,
                'amortizacion': amort_o,
            }
        )
        n += 1
    print(f'✓ Refinanciación proyección:     {n:>5}')

    # ═══ 11) CCPET INGRESOS 2027 ═════════════════════════════════════════
    ws = wb['CCPET Ingresos 2027']
    n = 0
    for r in range(5, ws.max_row + 1):
        cod = ws.cell(row=r, column=1).value
        if not cod: continue
        fte = ws.cell(row=r, column=2).value or ''
        desc = ws.cell(row=r, column=3).value or ''
        ppto = D(ws.cell(row=r, column=4).value)
        CCPETIngreso.objects.update_or_create(
            vigencia=2027, rubro_presupuestal=str(cod).strip(),
            fuente=str(fte or '').strip(),
            defaults={'descripcion': str(desc).strip()[:2000], 'presupuesto': ppto},
        )
        n += 1
    print(f'✓ CCPET Ingresos 2027:           {n:>5}')

    # ═══ 12) CCPET GASTOS 2027 ═══════════════════════════════════════════
    ws = wb['CCPET Gastos 2027']
    n = 0
    for r in range(5, ws.max_row + 1):
        cod = ws.cell(row=r, column=1).value
        if not cod: continue
        fte = ws.cell(row=r, column=2).value or ''
        desc = ws.cell(row=r, column=3).value or ''
        ppto = D(ws.cell(row=r, column=4).value)
        CCPETGasto.objects.update_or_create(
            vigencia=2027, rubro_presupuestal=str(cod).strip(),
            fuente=str(fte or '').strip(),
            defaults={'descripcion': str(desc).strip()[:2000], 'presupuesto': ppto},
        )
        n += 1
    print(f'✓ CCPET Gastos 2027:             {n:>5}')

    # ═══ 13) PROYECCIÓN RUBROS INGRESO 10 años (hoja 'Ingresos') ═══════
    ws = wb['Ingresos']
    n = 0
    # Headers en F1. Datos desde F2. 40 columnas.
    # cols: 1=Cod, 2=Fte, 3=NomFuente, 4=Desc, 5=Ej2024, 6=Ej2025, 7=Aforo2026,
    # 8=RecaudoYTD, 9=%PromHist, 10=ProyDic, 11=Metodo, 12-21=2027-2036
    for r in range(2, ws.max_row + 1):
        cod = ws.cell(row=r, column=1).value
        if not cod: continue
        fte = str(ws.cell(row=r, column=2).value or '').strip()
        ProyeccionRubroIngreso.objects.update_or_create(
            codigo_ccpet=str(cod).strip(), codigo_fuente=fte,
            defaults={
                'nombre_fuente': str(ws.cell(row=r, column=3).value or '')[:200],
                'descripcion': str(ws.cell(row=r, column=4).value or '')[:400],
                'ejec_2024': D(ws.cell(row=r, column=5).value),
                'ejec_2025': D(ws.cell(row=r, column=6).value),
                'aforo_2026': D(ws.cell(row=r, column=7).value),
                'recaudo_ytd_2026': D(ws.cell(row=r, column=8).value),
                'pct_prom_historico': D(ws.cell(row=r, column=9).value),
                'proyeccion_dic_2026': D(ws.cell(row=r, column=10).value),
                'metodo': str(ws.cell(row=r, column=11).value or '')[:100],
                'proy_2027': D(ws.cell(row=r, column=12).value),
                'proy_2028': D(ws.cell(row=r, column=13).value),
                'proy_2029': D(ws.cell(row=r, column=14).value),
                'proy_2030': D(ws.cell(row=r, column=15).value),
                'proy_2031': D(ws.cell(row=r, column=16).value),
                'proy_2032': D(ws.cell(row=r, column=17).value),
                'proy_2033': D(ws.cell(row=r, column=18).value),
                'proy_2034': D(ws.cell(row=r, column=19).value),
                'proy_2035': D(ws.cell(row=r, column=20).value),
                'proy_2036': D(ws.cell(row=r, column=21).value),
            },
        )
        n += 1
    print(f'✓ Proyección Rubros Ingreso:     {n:>5}')

    # ═══ 14) PROYECCIÓN RUBROS GASTO 10 años (hoja 'Gastos') ═══════════
    ws = wb['Gastos']
    n = 0
    # cols: 1=Cod, 2=Fte, 3=NomFuente, 4=Desc, 5=Categoria,
    # 6=Aprop2026, 7=CompMayo, 8=ProyDic, 9=Metodo, 10-19=2027-2036
    for r in range(2, ws.max_row + 1):
        cod = ws.cell(row=r, column=1).value
        if not cod: continue
        fte = str(ws.cell(row=r, column=2).value or '').strip()
        ProyeccionRubroGasto.objects.update_or_create(
            codigo_ccpet=str(cod).strip(), codigo_fuente=fte,
            defaults={
                'nombre_fuente': str(ws.cell(row=r, column=3).value or '')[:200],
                'descripcion': str(ws.cell(row=r, column=4).value or '')[:400],
                'categoria': str(ws.cell(row=r, column=5).value or '')[:100],
                'apropiacion_2026': D(ws.cell(row=r, column=6).value),
                'compromiso_mayo_2026': D(ws.cell(row=r, column=7).value),
                'proyeccion_dic_2026': D(ws.cell(row=r, column=8).value),
                'metodo': str(ws.cell(row=r, column=9).value or '')[:100],
                'proy_2027': D(ws.cell(row=r, column=10).value),
                'proy_2028': D(ws.cell(row=r, column=11).value),
                'proy_2029': D(ws.cell(row=r, column=12).value),
                'proy_2030': D(ws.cell(row=r, column=13).value),
                'proy_2031': D(ws.cell(row=r, column=14).value),
                'proy_2032': D(ws.cell(row=r, column=15).value),
                'proy_2033': D(ws.cell(row=r, column=16).value),
                'proy_2034': D(ws.cell(row=r, column=17).value),
                'proy_2035': D(ws.cell(row=r, column=18).value),
                'proy_2036': D(ws.cell(row=r, column=19).value),
            },
        )
        n += 1
    print(f'✓ Proyección Rubros Gasto:       {n:>5}')

    # ═══ 15) CARGA POAI PROYECTOS (BPIN) ═══════════════════════════════
    ws = wb['Carga POAI']
    n = 0
    for r in range(5, ws.max_row + 1):
        numero = ws.cell(row=r, column=1).value
        if numero is None: continue
        try: numero = int(numero)
        except: continue
        CargaPOAIProyecto.objects.update_or_create(
            numero=numero,
            defaults={
                'codigo_rubro': str(ws.cell(row=r, column=2).value or '')[:50],
                'codigo_fuente': str(ws.cell(row=r, column=3).value or '')[:10],
                'nombre_fuente': str(ws.cell(row=r, column=4).value or '')[:200],
                'proyecto_bpin': str(ws.cell(row=r, column=5).value or '')[:2000],
                'valor_poai_2027': D(ws.cell(row=r, column=6).value),
            },
        )
        n += 1
    print(f'✓ Carga POAI Proyectos:          {n:>5}')

    # ═══ 16) ICO PROYECCIÓN por actividad CIIU ═══════════════════════════
    ws = wb['ICO']
    n = 0
    for r in range(5, ws.max_row + 1):
        ciiu = ws.cell(row=r, column=1).value
        if not ciiu: continue
        ICOProyeccion.objects.update_or_create(
            codigo_ciiu=str(ciiu).strip(),
            defaults={
                'descripcion': str(ws.cell(row=r, column=2).value or '')[:400],
                'contribuyentes_2024': int(D(ws.cell(row=r, column=3).value)),
                'ico_liquidado_2024': D(ws.cell(row=r, column=4).value),
                'proy_2026': D(ws.cell(row=r, column=5).value),
                'proy_2027': D(ws.cell(row=r, column=6).value),
                'proy_2028': D(ws.cell(row=r, column=7).value),
                'proy_2029': D(ws.cell(row=r, column=8).value),
                'proy_2030': D(ws.cell(row=r, column=9).value),
                'proy_2031': D(ws.cell(row=r, column=10).value),
                'proy_2032': D(ws.cell(row=r, column=11).value),
                'proy_2033': D(ws.cell(row=r, column=12).value),
                'proy_2034': D(ws.cell(row=r, column=13).value),
                'proy_2035': D(ws.cell(row=r, column=14).value),
                'proy_2036': D(ws.cell(row=r, column=15).value),
            },
        )
        n += 1
    print(f'✓ ICO Proyección CIIU:           {n:>5}')

    # ═══ 17) PLANTA DETALLE por cargo (52 filas) ═══════════════════════
    ws = wb['Planta Detalle']
    n = 0
    for r in range(5, ws.max_row + 1):
        sec = ws.cell(row=r, column=1).value
        denom = ws.cell(row=r, column=3).value
        if not sec or not denom: continue
        PlantaDetalleCargo.objects.update_or_create(
            seccion=str(sec).strip()[:200],
            nivel=str(ws.cell(row=r, column=2).value or '')[:50],
            denominacion=str(denom).strip()[:200],
            codigo_cargo=str(ws.cell(row=r, column=4).value or '')[:20],
            defaults={
                'grado': str(ws.cell(row=r, column=5).value or '')[:10],
                'cantidad': int(D(ws.cell(row=r, column=6).value) or 1),
                'nombre_fuente': str(ws.cell(row=r, column=7).value or '')[:100],
                'crece_por': str(ws.cell(row=r, column=8).value or '')[:50],
                'asig_mensual_2026': D(ws.cell(row=r, column=9).value),
                'costo_anual_2026': D(ws.cell(row=r, column=10).value),
                'costo_2027': D(ws.cell(row=r, column=11).value),
                'costo_2028': D(ws.cell(row=r, column=12).value),
                'costo_2029': D(ws.cell(row=r, column=13).value),
                'costo_2030': D(ws.cell(row=r, column=14).value),
                'costo_2031': D(ws.cell(row=r, column=15).value),
                'costo_2032': D(ws.cell(row=r, column=16).value),
                'costo_2033': D(ws.cell(row=r, column=17).value),
                'costo_2034': D(ws.cell(row=r, column=18).value),
                'costo_2035': D(ws.cell(row=r, column=19).value),
                'costo_2036': D(ws.cell(row=r, column=20).value),
            },
        )
        n += 1
    print(f'✓ Planta Detalle Cargos:         {n:>5}')

    # ═══ 18) PARÁMETROS ANUALES PREDIAL ══════════════════════════════════
    ws = wb['Predial']
    # F4: años cols 5..25 impares
    anios_predial = []
    for c in range(5, 26):
        v = ws.cell(row=4, column=c).value
        if v and isinstance(v, (int, float)) and 2020 <= int(v) <= 2050:
            anios_predial.append((c, int(v)))
    n = 0
    for col, anio in anios_predial:
        eff_urb = D(ws.cell(row=5, column=col).value)
        eff_rur = D(ws.cell(row=6, column=col).value)
        aju = D(ws.cell(row=7, column=col).value)
        base_c = D(ws.cell(row=8, column=col).value)
        cart_urb = D(ws.cell(row=9, column=col).value)
        cart_rur = D(ws.cell(row=10, column=col).value) if ws.cell(row=10, column=col).value else Decimal('0.80')
        ParametroAnualPredial.objects.update_or_create(
            anio=anio, defaults={
                'pct_eficiencia_urbano': eff_urb,
                'pct_eficiencia_rural': eff_rur,
                'pct_ajuste_avaluo': aju,
                'pct_base_cartera': base_c,
                'pct_cartera_urbano': cart_urb,
                'pct_cartera_rural': cart_rur,
            },
        )
        n += 1
    print(f'✓ Parámetros Anuales Predial:    {n:>5}')

    # ═══ 19) PARÁMETROS ANUALES PLANTA (Costo Planta Personal) ═══════════
    ws = wb['Costo Planta Personal']
    anios_planta = []
    for c in range(3, 14):
        v = ws.cell(row=5, column=c).value
        if v and isinstance(v, (int, float)) and 2020 <= int(v) <= 2050:
            anios_planta.append((c, int(v)))
    n = 0
    for col, anio in anios_planta:
        ipc = D(ws.cell(row=6, column=col).value)
        ipc_ref = D(ws.cell(row=7, column=col).value)
        prod = D(ws.cell(row=8, column=col).value)
        pts = D(ws.cell(row=9, column=col).value)
        if ipc == 0 and ipc_ref == 0 and prod == 0:
            continue
        ParametroAnualPlanta.objects.update_or_create(
            anio=anio, defaults={
                'ipc_esperado': ipc,
                'ipc_ref_mfmp': ipc_ref,
                'indice_productividad': prod,
                'puntos_salariales_sindicales': pts,
            },
        )
        n += 1
    print(f'✓ Parámetros Anuales Planta:     {n:>5}')

    # ═══ 20) BASE ESTAMPILLAS por año ═════════════════════════════════════
    ws = wb['Estampillas']
    anios_est = []
    for c in range(3, 14):
        v = ws.cell(row=4, column=c).value
        if v and isinstance(v, (int, float)) and 2020 <= int(v) <= 2050:
            anios_est.append((c, int(v)))
    n = 0
    for col, anio in anios_est:
        BaseEstampillasAnual.objects.update_or_create(
            anio=anio, defaults={
                'valor_total_poai': D(ws.cell(row=5, column=col).value),
                'gasto_apropiado_sev_ppto': D(ws.cell(row=6, column=col).value),
                'saldo_neto_ppto': D(ws.cell(row=7, column=col).value),
                'presupuesto_sgr': D(ws.cell(row=8, column=col).value),
                'gasto_apropiado_sev_sgr': D(ws.cell(row=9, column=col).value),
                'saldo_neto_sgr': D(ws.cell(row=10, column=col).value),
                'reservas_ppto_nc': D(ws.cell(row=11, column=col).value),
                'cuentas_por_pagar_nc': D(ws.cell(row=12, column=col).value),
                'superavit_fiscal': D(ws.cell(row=13, column=col).value),
            },
        )
        n += 1
    print(f'✓ Base Estampillas Anual:        {n:>5}')

    # ═══ Sincronizar ICLD calculado de ParametrosSistema desde Plan Financiero
    # (usuario pidió: ICLD 2027 debe venir del Plan Financiero proyectado 2027-2036)
    p = ParametrosSistema.objects.filter(activo=True).first()
    if p:
        # ICLD 2027 total = fuente "RECURSOS PROPIOS" (código 1) del ICLDProyectado
        icld_2027 = ICLDProyectado.objects.filter(
            fuente__codigo='1', anio=p.vigencia,
        ).first()
        if icld_2027 and icld_2027.valor_bruto > 0:
            p.icld_calculado = icld_2027.valor_bruto
            p.save(update_fields=['icld_calculado'])
            print(f'✓ ICLD ParametrosSistema (2027): ${p.icld_calculado:,.0f}')


if __name__ == '__main__':
    ruta = sys.argv[1] if len(sys.argv) > 1 else '/Users/jorgebinkio/Downloads/Sistema_MFMP_PuertoLopez_v75.xlsx'
    if not os.path.isfile(ruta):
        print(f'No existe: {ruta}')
        sys.exit(1)
    print(f'Importando: {ruta}\n')
    importar(ruta)
    print('\n=== IMPORT COMPLETADO ===')
