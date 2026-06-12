"""Importa al sistema los datos del Excel CALCULO_GASTOS.xlsx:

1. Actualiza parametros Anexo 6 con valores del Excel nuevo:
   - SMLMV = $2,200,000 (proyectado 2027)
   - Valor sesion concejal cat 5 = $328,554.44
   - Progresion SMLV Personeria cat 5 actualizada (2026=210, 2027=220, ...)

2. Importa COSTO PERSONAL de las 5 hojas de planta a CostoPersonal:
   - Concejo (1 cargo)
   - Personeria (2 cargos)
   - Admon central (33 cargos) -> seccion 03
   - Comisaria (3 cargos) -> seccion 03
   - Salud 7% (5 cargos) -> seccion 04

3. Importa COSTO PENSIONADOS (hoja Costo Pensionados, 7 pensionados):
   - Total mesadas anual: $241,208,230 -> seccion 09 Pensiones

4. Importa amortizacion de DEUDA PUBLICA (hoja Deuda Publica):
   - Pagare ITO $25,000M, 9.9% anual, 10 anos, gracia 2 anos
   - ~40 cuotas trimestrales con K, I, K+I, Saldo, Fecha

5. Asigna metodo CPS a los rubros 2.1.1.x de cada seccion (CONCEJO,
   PERSONERIA, DESPACHO, SALUD) para que se autocalculen desde CostoPersonal.

6. Asigna metodo PEN a rubros de pensionados (seccion 09).

7. Recalcula todo.

Idempotente: limpia CostoPersonal y AmortizacionPagare antes de importar
para evitar duplicados. Tu data se conserva (backup ya hecho aparte).
"""
import os
from decimal import Decimal
from datetime import date
import django

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'presupuesto_project.settings')
django.setup()

from openpyxl import load_workbook
from django.db import transaction

from core.models import (
    ParametrosSistema, TablaConcejoPersoneria, PersoneriaSMLVProgresion,
)
from gastos.models import (
    RubroGasto, CostoPersonal, SeccionGasto, ContratoCredito, PagareCredito,
    AmortizacionPagare,
)
from gastos.utils import recalcular_rubros_metodo


EXCEL_PATH = '/tmp/calc_gastos.xlsx'


# Map hoja Excel -> codigo SeccionGasto
HOJAS_A_SECCION = [
    ('Costo planta Concejo',       '01'),
    ('Costo planta Personeria',    '02'),
    ('Costo planta admon central', '03'),
    ('Costo planta Comisaria',     '03'),
    ('Costo planta salud 7%',      '04'),
]


def update_anexo6():
    print('\n========== 1. UPDATE ANEXO 6 ==========')
    p = (ParametrosSistema.objects.filter(activo=True).first()
         or ParametrosSistema.objects.order_by('-vigencia').first())
    v = p.vigencia
    print(f'Vigencia activa: {v}')

    p.valor_smlmv = Decimal('2200000.00')
    p.save(update_fields=['valor_smlmv'])
    print(f'SMLMV actualizado a ${p.valor_smlmv:,.0f}')

    cat5 = TablaConcejoPersoneria.objects.filter(categoria=5).first()
    if cat5:
        cat5.valor_sesion_concejal = Decimal('328554.44')
        cat5.save(update_fields=['valor_sesion_concejal'])
        print(f'Valor sesion cat 5: ${cat5.valor_sesion_concejal:,.2f}')

    # Progresion SMLV cat 5 segun Excel Variables Macro:
    # 2026=210, 2027=220, 2028=230, 2029=240, 2030-2035=240
    progresion_cat5 = {2026: 210, 2027: 220, 2028: 230, 2029: 240,
                       2030: 240, 2031: 240, 2032: 240, 2033: 240,
                       2034: 240, 2035: 240}
    for vig, smlv in progresion_cat5.items():
        PersoneriaSMLVProgresion.objects.update_or_create(
            vigencia=vig, categoria=5, defaults={'smlv': smlv})
    print(f'Progresion cat 5 actualizada para 2026-2035')

    # Verificacion
    tc = cat5.calcular_transferencia_concejo(
        p.icld_calculado or Decimal('0'),
        p.valor_smlmv,
        p.pct_icld_adicional_concejo or Decimal('0'))
    tp = cat5.calcular_transferencia_personeria(
        v, p.icld_calculado or Decimal('0'), p.valor_smlmv)
    print(f'  -> Concejo {v}:    ${tc:,.0f}')
    print(f'  -> Personeria {v}: ${tp:,.0f}')


def _safe_decimal(v):
    if v is None:
        return Decimal('0')
    try:
        return Decimal(str(v))
    except Exception:
        return Decimal('0')


def importar_planta(v):
    print('\n========== 2. IMPORTAR PLANTA ==========')
    wb = load_workbook(EXCEL_PATH, data_only=True)

    # Limpia las plantas activas (no pensionados) de la vigencia para evitar dupes
    eliminados = CostoPersonal.objects.filter(vigencia=v, es_pensionado=False).delete()
    print(f'CostoPersonal activos previos eliminados: {eliminados[0]}')

    total_importado = 0
    for hoja, cod_sec in HOJAS_A_SECCION:
        ws = wb[hoja]
        sec = SeccionGasto.objects.filter(codigo=cod_sec).first()
        if not sec:
            print(f'  ! Seccion {cod_sec} no encontrada para hoja {hoja}')
            continue

        n_hoja = 0
        for r in range(10, ws.max_row + 1):
            nivel = ws.cell(row=r, column=1).value or ''
            cargo = ws.cell(row=r, column=2).value or ''
            cod = ws.cell(row=r, column=3).value
            grado = ws.cell(row=r, column=4).value
            asign_2027 = ws.cell(row=r, column=6).value  # F
            cantidad = ws.cell(row=r, column=7).value     # G
            gran_total = ws.cell(row=r, column=48).value  # AV

            cargo_s = str(cargo).strip() if cargo else ''
            nivel_s = str(nivel).strip().upper() if nivel else ''

            # Skip filas vacias o totales
            if not cargo_s or 'TOTAL' in cargo_s.upper() or 'TOTAL' in nivel_s:
                continue
            if 'PERSONAL DE' in nivel_s or 'RESUMEN' in nivel_s:
                continue
            if not cantidad or cantidad == 0:
                continue

            # Datos del Excel son por toda la planta (subtotales).
            # Modelo guarda valor unitario y al consultar multiplica por cantidad.
            # Para que la propiedad costo_total_anual de el valor correcto,
            # usamos el override.
            try:
                cantidad_int = int(cantidad)
            except Exception:
                continue

            CostoPersonal.objects.create(
                vigencia=v,
                seccion=sec,
                cargo=cargo_s[:200],
                grado=str(grado or '')[:20],
                cantidad=cantidad_int,
                salario_basico=_safe_decimal(asign_2027),
                # Override con el gran total del Excel
                costo_total_anual_override=_safe_decimal(gran_total),
                es_pensionado=False,
                observaciones=f'Codigo cargo: {cod or ""}; Importado desde {hoja}',
            )
            n_hoja += 1
            total_importado += 1
        print(f'  {hoja} -> seccion {cod_sec} {sec.nombre[:35]}: {n_hoja} cargos importados')

    # Verificacion: total por seccion
    print('\n  Totales por seccion (suma de costo_total_anual):')
    for sec in SeccionGasto.objects.all().order_by('codigo'):
        total = sum(cp.costo_total_anual for cp in CostoPersonal.objects.filter(
            vigencia=v, seccion=sec, es_pensionado=False))
        if total > 0:
            print(f'    Sec {sec.codigo} {sec.nombre[:30]:30}: ${total:,.0f}')

    return total_importado


def importar_pensionados(v):
    print('\n========== 3. IMPORTAR PENSIONADOS ==========')
    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['Costo Pensionados']

    # Limpia pensionados previos
    eliminados = CostoPersonal.objects.filter(vigencia=v, es_pensionado=True).delete()
    print(f'Pensionados previos eliminados: {eliminados[0]}')

    sec09 = SeccionGasto.objects.filter(codigo='09').first()
    if not sec09:
        print('  ! Seccion 09 no encontrada')
        return 0

    # Filas 3-9 son los 7 pensionados (B=cc pensionado, C=nombre, D=cc beneficiario,
    # E=beneficiario, F=mesada).
    # Total anual = mesada × 14 (incluye prima junio + prima diciembre)
    incremento = Decimal('1.07')  # Excel: 7% de incremento aprox
    n = 0
    for r in range(3, 10):
        cc = ws.cell(row=r, column=2).value
        nombre = ws.cell(row=r, column=3).value
        beneficiario = ws.cell(row=r, column=5).value
        mesada = ws.cell(row=r, column=6).value
        if not nombre or not mesada:
            continue
        mesada_d = _safe_decimal(mesada)
        # Aplicar incremento como en el Excel y multiplicar por 14
        mesada_incr = (mesada_d * incremento).quantize(Decimal('0.01'))
        total_anual = (mesada_incr * 14).quantize(Decimal('0.01'))

        obs = f'CC {cc}; Beneficiario: {beneficiario or "N/A"}; Mesada {mesada_d}'
        CostoPersonal.objects.create(
            vigencia=v,
            seccion=sec09,
            cargo=f'PENSIONADO {str(nombre).strip()[:120]}',
            grado='',
            cantidad=1,
            salario_basico=mesada_incr,
            costo_total_anual_override=total_anual,
            es_pensionado=True,
            observaciones=obs,
        )
        n += 1
        print(f'  Pensionado {n}: {str(nombre)[:40]:40} mesada=${mesada_d:,.0f} anual=${total_anual:,.0f}')

    total_pens = sum(cp.costo_total_anual for cp in
                     CostoPersonal.objects.filter(vigencia=v, es_pensionado=True))
    print(f'  TOTAL PENSIONADOS ANUAL: ${total_pens:,.0f}')
    return n


def importar_amortizacion(v):
    print('\n========== 4. IMPORTAR AMORTIZACION DEUDA ==========')
    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['Deuda Publica']

    pagare = PagareCredito.objects.first()
    if not pagare:
        print('  ! No hay PagareCredito en BD. Saltando amortizacion.')
        return 0

    eliminados = AmortizacionPagare.objects.filter(pagare=pagare).delete()
    print(f'Amortizaciones previas eliminadas: {eliminados[0]}')

    # AmortizacionPagare es POR VIGENCIA (no por cuota). Agrupamos las cuotas
    # trimestrales por año fiscal y guardamos una sola fila por año.
    # TCR se calcula como (intereses_sin_TCR) * factor TCR donde el factor
    # del Excel es 0.921 sobre el interés sin TCR para totalizar 1.921x.
    # Pero la columna C ya da intereses sin TCR puros. El sistema guarda
    # intereses (sin TCR) e intereses_tcr (solo el componente TCR adicional).
    TCR_FACTOR = Decimal('0.921')

    # Agrupar
    grupos = {}  # year -> {'capital': X, 'intereses': Y, 'cuotas': N}
    for r in range(17, ws.max_row + 1):
        cuota = ws.cell(row=r, column=1).value
        capital = ws.cell(row=r, column=2).value
        intereses = ws.cell(row=r, column=3).value
        fecha = ws.cell(row=r, column=6).value
        if not cuota or not fecha or not hasattr(fecha, 'year'):
            continue
        try:
            int(cuota)
        except Exception:
            continue
        year = fecha.year
        g = grupos.setdefault(year, {'capital': Decimal('0'),
                                       'intereses': Decimal('0'),
                                       'cuotas': 0})
        g['capital'] += _safe_decimal(capital)
        g['intereses'] += _safe_decimal(intereses)
        g['cuotas'] += 1

    n = 0
    for year in sorted(grupos.keys()):
        g = grupos[year]
        tcr = (g['intereses'] * TCR_FACTOR).quantize(Decimal('0.01'))
        AmortizacionPagare.objects.create(
            pagare=pagare,
            vigencia_pago=year,
            capital_principal=g['capital'],
            intereses=g['intereses'],
            intereses_tcr=tcr,
        )
        n += 1
        total = g['capital'] + g['intereses'] + tcr
        print(f'  {year}: {g["cuotas"]} cuotas  K=${g["capital"]:,.0f}  '
              f'I=${g["intereses"]:,.0f}  TCR=${tcr:,.0f}  Total=${total:,.0f}')
    return n


def asignar_metodos_cps_pen(v):
    print('\n========== 5. ASIGNAR METODOS CPS / PEN ==========')
    # CPS: todos los rubros bajo "2.1.1.01 Planta de personal permanente"
    # de cada seccion (excepto pensionados)
    # PEN: el rubro de mesadas pensionales (cualquiera bajo seccion 09 con 'pension' o 'mesada')

    secciones_cps = ['01', '02', '03', '04']
    for cod_sec in secciones_cps:
        sec = SeccionGasto.objects.filter(codigo=cod_sec).first()
        if not sec:
            continue
        # Cubre todos los rubros HOJA bajo 2.1.1 (gastos de personal) de la seccion
        rubros = RubroGasto.objects.filter(
            vigencia=v, seccion=sec, es_titulo=False,
            codigo__startswith='2.1.1')
        # Asignamos CPS al rubro 2.1.1.01 si existe, y dejamos los hijos en MAN
        # (porque el modelo CostoPersonal da el total, no el desglose).
        rubro_planta = rubros.filter(codigo='2.1.1.01').first()
        if rubro_planta:
            if rubro_planta.metodo_calculo != 'CPS':
                rubro_planta.metodo_calculo = 'CPS'
                rubro_planta.save(update_fields=['metodo_calculo'])
                print(f'  Sec {cod_sec}: rubro 2.1.1.01 -> CPS')

    # PEN: para seccion 09 Fondo Pensiones
    sec09 = SeccionGasto.objects.filter(codigo='09').first()
    if sec09:
        rubros_pens = RubroGasto.objects.filter(vigencia=v, seccion=sec09, es_titulo=False)
        # Buscar el rubro de mesadas pensionales (descripcion contiene "pension" o "mesada")
        from django.db.models import Q
        rubro_pen = rubros_pens.filter(
            Q(descripcion__icontains='pension') | Q(descripcion__icontains='mesada')
        ).first()
        if rubro_pen:
            if rubro_pen.metodo_calculo != 'PEN':
                rubro_pen.metodo_calculo = 'PEN'
                rubro_pen.save(update_fields=['metodo_calculo'])
                print(f'  Sec 09: {rubro_pen.codigo} {rubro_pen.descripcion[:50]} -> PEN')
        else:
            print(f'  Sec 09: no se encontro rubro de mesadas/pensiones')


def recalcular_y_verificar(v):
    print('\n========== 6. RECALCULAR Y VERIFICAR ==========')
    resumen = recalcular_rubros_metodo(v)
    for m, info in resumen.items():
        if info['rubros']:
            print(f'  {m}: {info["rubros"]} rubros, ${info["total_aplicado"]:,.0f}')


@transaction.atomic
def run():
    p = (ParametrosSistema.objects.filter(activo=True).first()
         or ParametrosSistema.objects.order_by('-vigencia').first())
    v = p.vigencia
    print(f'=== IMPORTANDO CALCULO_GASTOS.xlsx para vigencia {v} ===')

    update_anexo6()
    n_planta = importar_planta(v)
    n_pen = importar_pensionados(v)
    n_amort = importar_amortizacion(v)
    asignar_metodos_cps_pen(v)
    recalcular_y_verificar(v)

    print('\n=== RESUMEN ===')
    print(f'  Cargos importados:    {n_planta}')
    print(f'  Pensionados:          {n_pen}')
    print(f'  Cuotas amortizacion:  {n_amort}')


if __name__ == '__main__':
    run()
