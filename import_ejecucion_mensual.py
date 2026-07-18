"""Importa Ejecución Mensualizada de Ingresos 2024 y 2025 a la BD.

Uso:
    python import_ejecucion_mensual.py

Después de importar, se ejecuta automáticamente `actualizar_pct_prom_historico()`
que calcula el % promedio de recaudo hasta el mes de corte (junio por defecto)
usando los datos históricos 2024 + 2025, y con ese % + el recaudo YTD del año
en curso (aforo/proyeccion_dic_2026 de ProyeccionRubroIngreso), estima la
proyección a diciembre del año en curso.
"""
import os, sys, warnings
warnings.filterwarnings('ignore')
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'presupuesto_project.settings')
import django; django.setup()
from decimal import Decimal, InvalidOperation
from openpyxl import load_workbook
from django.db.models import Sum
from core.models import EjecucionMensualIngreso, ProyeccionRubroIngreso


def D(v):
    if v is None or v == '':
        return Decimal('0')
    try:
        return Decimal(str(v))
    except (InvalidOperation, ValueError):
        return Decimal('0')


def _codigo_normalizado(raw):
    """Normaliza códigos '03: 1.1.01.01.014.01' → '03.1.1.01.01.014.01'."""
    s = str(raw or '').strip()
    return s.replace(':', '.').replace(' ', '').rstrip('.')


def _fuente_extraida(raw):
    """De 'TOTAL RUBRO' o '1 - RECURSOS PROPIOS' extrae '' o '1'."""
    s = str(raw or '').strip()
    if not s or 'TOTAL RUBRO' in s.upper():
        return ''
    # Ej: '1 - RECURSOS PROPIOS' → '1'
    parts = s.split(' - ', 1)
    if parts:
        return parts[0].strip()
    return s


def importar_anio(xlsx_path, anio):
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb['Hoja1']
    # Header en fila 6, datos desde fila 7
    n = 0
    for r in range(7, ws.max_row + 1):
        cod_raw = ws.cell(row=r, column=1).value
        tipo = str(ws.cell(row=r, column=2).value or '').strip()
        if not cod_raw or tipo == 'MAYOR':
            continue
        fuente_raw = ws.cell(row=r, column=4).value
        codigo = _codigo_normalizado(cod_raw)
        fuente = _fuente_extraida(fuente_raw)
        # Solo importar filas con fuente específica (excluir 'TOTAL RUBRO')
        if not fuente:
            continue
        EjecucionMensualIngreso.objects.update_or_create(
            anio=anio, codigo_ccpet=codigo, codigo_fuente=fuente,
            defaults={
                'descripcion': str(ws.cell(row=r, column=3).value or '')[:400],
                'apropiacion_definitiva': D(ws.cell(row=r, column=8).value),
                'ene': D(ws.cell(row=r, column=10).value),
                'feb': D(ws.cell(row=r, column=11).value),
                'mar': D(ws.cell(row=r, column=12).value),
                'abr': D(ws.cell(row=r, column=13).value),
                'may': D(ws.cell(row=r, column=14).value),
                'jun': D(ws.cell(row=r, column=15).value),
                'jul': D(ws.cell(row=r, column=16).value),
                'ago': D(ws.cell(row=r, column=17).value),
                'sep': D(ws.cell(row=r, column=18).value),
                'oct': D(ws.cell(row=r, column=19).value),
                'nov': D(ws.cell(row=r, column=20).value),
                'dic': D(ws.cell(row=r, column=21).value),
            }
        )
        n += 1
    print(f'✓ Ejecución {anio}: {n} rubros importados')


def actualizar_pct_prom_historico(mes_corte=6):
    """Actualiza el % promedio histórico y la proyección diciembre del año
    en curso (vigencia activa - 1, generalmente 2026 mientras se prepara 2027)
    para cada ProyeccionRubroIngreso.

    Regla:
        pct_prom = promedio(pct_2024, pct_2025)
        donde pct_año = SUM(ene..mes_corte año) / SUM(ene..dic año)

        proy_dic_actual = recaudo_ytd_2026 / pct_prom
        (si pct_prom = 0, se conserva el aforo)

    Excepciones NO actualizadas (siguen con su lógica específica):
        - Predial (ya tiene cálculo con avalúo × tarifa × eficiencia)
        - ICA (con contribuyentes)
        - Estampillas (base × tarifa)
        - Impuesto Transporte Oleoductos - ITO (promedio 3 años)
    """
    # Los rubros que ya tienen su propio método de cálculo NO usan este:
    # códigos por prefijo CCPET (más confiable que descripción):
    #   03.1.1.01.01.200 → Predial
    #   03.1.1.01.02.200 → Industria y Comercio (ICA)
    #   03.1.1.01.02.300 → Estampillas
    #   03.1.1.01.02.214 → Transporte Oleoductos (ITO)
    PREFIJOS_EXCLUIDOS = (
        '03.1.1.01.01.200',   # Predial
        '03.1.1.01.02.200',   # ICA
        '03.1.1.01.02.300',   # Estampillas
        '03.1.1.01.02.214',   # ITO
    )

    def es_excluido(rubro):
        cod = rubro.codigo_ccpet or ''
        return any(cod.startswith(p) for p in PREFIJOS_EXCLUIDOS)

    campos = ['ene','feb','mar','abr','may','jun','jul','ago','sep','oct','nov','dic']

    actualizados = 0
    no_encontrados = 0
    for pri in ProyeccionRubroIngreso.objects.all():
        if es_excluido(pri):
            continue
        # Buscar ejecución 2024 y 2025 del mismo rubro+fuente
        eje24 = EjecucionMensualIngreso.objects.filter(
            anio=2024, codigo_ccpet=pri.codigo_ccpet, codigo_fuente=pri.codigo_fuente
        ).first()
        eje25 = EjecucionMensualIngreso.objects.filter(
            anio=2025, codigo_ccpet=pri.codigo_ccpet, codigo_fuente=pri.codigo_fuente
        ).first()

        pcts = []
        for eje in (eje24, eje25):
            if not eje:
                continue
            total_dic = eje.total
            if total_dic <= 0:
                continue
            acum = sum(getattr(eje, campos[i]) for i in range(mes_corte))
            pct = acum / total_dic
            pcts.append(pct)

        if not pcts:
            no_encontrados += 1
            continue

        pct_prom = sum(pcts) / Decimal(str(len(pcts)))
        # Actualizar solo si hay pct válido > 0
        if pct_prom <= 0:
            continue
        pri.pct_prom_historico = pct_prom
        # Proyectar dic 2026: recaudo_ytd_2026 / pct_prom
        if pri.recaudo_ytd_2026 and pri.recaudo_ytd_2026 > 0:
            proy_nueva = pri.recaudo_ytd_2026 / pct_prom
            # Sanity: no aceptar proyecciones > 5x el aforo (outliers)
            aforo = pri.aforo_2026 or Decimal('0')
            if aforo > 0 and proy_nueva > aforo * 5:
                proy_nueva = aforo
            pri.proyeccion_dic_2026 = proy_nueva
        pri.save(update_fields=['pct_prom_historico', 'proyeccion_dic_2026'])
        actualizados += 1

    print(f'✓ pct_prom_historico actualizado: {actualizados} rubros')
    print(f'  Sin datos históricos:            {no_encontrados}')


if __name__ == '__main__':
    r24 = '/Users/jorgebinkio/Downloads/Ejecucion Mensualizada de ingresos 2024.xlsx'
    r25 = '/Users/jorgebinkio/Downloads/Ejecucion ingresos mensualizada 2025.xlsx'
    if len(sys.argv) > 1:
        r24 = sys.argv[1]
    if len(sys.argv) > 2:
        r25 = sys.argv[2]
    print('Importando ejecución mensualizada:')
    importar_anio(r24, 2024)
    importar_anio(r25, 2025)
    print('\nCalculando % promedio histórico y proyección Dic 2026:')
    actualizar_pct_prom_historico(mes_corte=6)
    print('\n=== IMPORT COMPLETADO ===')
