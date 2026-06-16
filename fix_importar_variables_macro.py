"""Importa Variables Macro del Excel y actualiza calculos dependientes.

Lee el Excel VARIABLES_MACRO.xlsx (hoja Variables Macro) y carga:
  - SMLV (con % anual)
  - IPC
  - PIB
  - PETROLEO
  - DTF
  - TRM

Luego sincroniza ParametrosSistema.valor_smlmv y tasa_ipc con el ano
vigente activo, y recalcula todos los rubros.

Idempotente: update_or_create por (anio, tipo).
"""
import os
from decimal import Decimal
import django

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'presupuesto_project.settings')
django.setup()

from openpyxl import load_workbook
from core.models import (
    VariableMacro, ParametrosSistema, get_smlv, get_ipc,
)


EXCEL_PATH = '/tmp/var_macro.xlsx'

# Mapeo (col_anio, col_valor, col_pct, tipo)
COLUMNAS = [
    ('B', 'C', 'D', 'PIB'),       # PIB anio, valor, % crec
    ('F', 'G', 'H', 'SMLV'),
    ('J', 'K', None, 'IPC'),      # IPC: K es el % directo, no hay columna de valor adicional
    ('M', 'N', None, 'PETROLEO'),
    ('P', 'Q', None, 'DTF'),
    ('S', 'T', None, 'TRM'),
]
START_ROW = 4
END_ROW = 31  # Excel tiene datos hasta 2036


def safe_num(v):
    if v is None:
        return None
    try:
        return Decimal(str(v))
    except Exception:
        return None


def es_proyectado(anio, tipo):
    # Datos historicos hasta 2025, proyecciones desde 2026
    return anio >= 2026


def run():
    print('========== IMPORTAR VARIABLES MACRO ==========')
    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['Variables Macro']

    creados = 0
    actualizados = 0

    for col_anio, col_valor, col_pct, tipo in COLUMNAS:
        print(f'\n>> {tipo}')
        for r in range(START_ROW, END_ROW + 1):
            anio_val = ws[f'{col_anio}{r}'].value
            if not anio_val:
                continue
            try:
                anio = int(anio_val)
            except Exception:
                continue
            if anio < 2010 or anio > 2040:
                continue

            valor = safe_num(ws[f'{col_valor}{r}'].value)
            pct = safe_num(ws[f'{col_pct}{r}'].value) if col_pct else Decimal('0')

            if tipo == 'IPC':
                # Para IPC, la columna K contiene el % directamente.
                # Guardamos en ambos campos para que get_ipc() lo encuentre.
                valor_final = valor or Decimal('0')
                pct_final = valor or Decimal('0')
            else:
                valor_final = valor or Decimal('0')
                pct_final = pct or Decimal('0')

            obj, created = VariableMacro.objects.update_or_create(
                anio=anio, tipo=tipo,
                defaults={
                    'valor': valor_final,
                    'pct_anual': pct_final,
                    'es_proyectado': es_proyectado(anio, tipo),
                })
            if created:
                creados += 1
            else:
                actualizados += 1
            print(f'  {anio} valor={valor_final}  pct={pct_final}  proyec={es_proyectado(anio, tipo)}')

    print(f'\n  Total: {creados} creados, {actualizados} actualizados')

    # Sincronizar ParametrosSistema con valores del ano vigente
    p = ParametrosSistema.objects.filter(activo=True).first()
    if p:
        smlv_vig = get_smlv(p.vigencia)
        ipc_vig = get_ipc(p.vigencia)
        print(f'\n========== SINCRONIZAR PARAMETROS ==========')
        print(f'Vigencia activa: {p.vigencia}')
        print(f'SMLV del año:    ${smlv_vig:,.0f}')
        print(f'IPC del año:     {ipc_vig}')
        if smlv_vig:
            p.valor_smlmv = smlv_vig
        if ipc_vig:
            p.tasa_ipc = ipc_vig
        p.save(update_fields=['valor_smlmv', 'tasa_ipc'])

        # Recalcular
        from ingresos.utils import calcular_todos_ingresos
        from gastos.utils import recalcular_rubros_metodo
        from ingresos.models import RubroIngreso
        from gastos.models import RubroGasto
        from core.models import TablaConcejoPersoneria
        calcular_todos_ingresos(p.vigencia)
        resumen = recalcular_rubros_metodo(p.vigencia)
        for t in RubroIngreso.objects.filter(vigencia=p.vigencia, es_titulo=True).order_by('-nivel'):
            t.calcular_hijos()
        for t in RubroGasto.objects.filter(vigencia=p.vigencia, es_titulo=True).order_by('-nivel'):
            t.calcular_hijos()

        # Verificar Personeria con el nuevo SMLV
        tabla = TablaConcejoPersoneria.objects.filter(categoria=p.categoria_municipio).first()
        if tabla:
            tp = tabla.calcular_transferencia_personeria(p.vigencia, p.icld_calculado, p.valor_smlmv)
            print(f'\nPersoneria cat {p.categoria_municipio} con SMLV ${smlv_vig:,.0f}: ${tp:,.0f}')

        for m, info in resumen.items():
            if info['rubros']:
                print(f'  {m}: {info["rubros"]} rubros, ${info["total_aplicado"]:,.0f}')


if __name__ == '__main__':
    run()
